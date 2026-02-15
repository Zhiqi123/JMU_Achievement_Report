# -*- coding: utf-8 -*-
"""
达成度报告生成器 - 核心处理逻辑
"""

import re
from collections import Counter
from typing import Callable, Optional

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import RichTextProperties, Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.utils import get_column_letter

from .config import Config


class AchievementProcessor:
    """达成度报告处理器"""

    def __init__(self, config: Config = None):
        self.config = config or Config()
        self._progress_callback: Optional[Callable[[str, int], None]] = None

    def set_progress_callback(self, callback: Callable[[str, int], None]):
        """设置进度回调函数

        Args:
            callback: 回调函数，参数为 (消息, 进度百分比)
        """
        self._progress_callback = callback

    def _report_progress(self, message: str, percent: int):
        """报告进度"""
        if self._progress_callback:
            self._progress_callback(message, percent)

    def extract_students_from_grades(self, grades_file: str) -> tuple[list[dict], list[str]]:
        """从成绩文件中提取所有学生数据（动态识别列结构）

        Returns:
            (学生数据列表, 警告信息列表)
        """
        self._report_progress("正在读取成绩文件...", 5)

        # 检查文件是否存在
        import os
        if not os.path.exists(grades_file):
            raise FileNotFoundError(f"文件不存在: {grades_file}")

        # 尝试读取Excel文件
        try:
            xl = pd.ExcelFile(grades_file)
        except PermissionError:
            raise PermissionError(f"无法读取文件，可能被其他程序占用: {os.path.basename(grades_file)}")
        except Exception as e:
            raise ValueError(f"无法读取Excel文件，文件可能已损坏: {str(e)}")

        all_students = []
        warnings = []  # 收集警告信息
        multi_col_sheets = 0  # 统计多列并排的工作表数量
        # 如果只有一个工作表，即使是Sheet1也要处理
        sheets_to_process = xl.sheet_names if len(xl.sheet_names) == 1 else [s for s in xl.sheet_names if s != 'Sheet1']
        total_sheets = len(sheets_to_process)
        processed_sheets = 0

        # 辅助函数（定义在循环外部，避免重复创建）
        def is_valid_student_id(sid: str) -> bool:
            """判断是否为有效学号
            放宽条件：长度>=5，且数字占比>=80%（允许少量字母）
            """
            if not sid or len(sid) < 5:
                return False
            digit_count = sum(1 for c in sid if c.isdigit())
            return digit_count / len(sid) >= 0.8

        def is_empty(val) -> bool:
            """判断值是否为空"""
            if val is None:
                return True
            try:
                if pd.isna(val):
                    return True
            except (ValueError, TypeError):
                pass
            return str(val).strip() == ''

        for sheet in sheets_to_process:

            df = pd.read_excel(xl, sheet_name=sheet, header=None)

            # ===== 1. 动态查找行政班信息 =====
            # 支持多种位置和格式：
            # - 顶部/底部: "行政班：XXX" / "班级：XXX" 格式
            # - 列数据: 列头为"班级"/"行政班"，每行有各自的班级
            # - 工作表名称: 如 "9007851-0001_音乐2212" 提取 "音乐2212"
            class_name = None  # 统一班级名（从顶部/底部/工作表名称提取）

            # 1.1 在整个表格中搜索 "行政班：XXX" 或 "班级：XXX" 格式
            for i in range(len(df)):
                for j in range(min(5, len(df.columns))):
                    cell_value = str(df.iloc[i, j]).strip() if pd.notna(df.iloc[i, j]) else ''
                    if not cell_value:
                        continue

                    # 匹配 "行政班：XXX" 或 "班级：XXX" 格式
                    match = re.search(r'(?:行政班|班级)[：:\s]\s*([^\s(（]+)', cell_value)
                    if match:
                        class_name = match.group(1).strip()
                        break
                if class_name:
                    break

            # 1.2 如果未找到，尝试从工作表名称提取班级信息
            # 例如 "9007851-0001_音乐2212" -> "音乐2212"
            # 排除明显不是班级名的关键词
            exclude_keywords = ['达成度报告', '成绩单', '成绩', '总评', '期末', '平时', '报告', '统计']
            if not class_name:
                sheet_match = re.search(r'_([^\d_][^_]+)$', sheet)
                if sheet_match:
                    candidate = sheet_match.group(1).strip()
                    if not any(kw in candidate for kw in exclude_keywords):
                        class_name = candidate

            # 1.3 如果仍未找到，尝试从文件名提取班级信息
            # 例如 "2022级计算机1班成绩单.xlsx" -> "计算机1班"
            # 或 "软件工程2301_成绩.xlsx" -> "软件工程2301"
            if not class_name:
                filename = os.path.basename(grades_file)
                filename_no_ext = os.path.splitext(filename)[0]
                # 排除明显不是班级名的关键词
                exclude_keywords = ['达成度报告', '成绩单', '成绩', '总评', '期末', '平时', '报告']
                # 尝试匹配常见班级格式（避免匹配日期如2023-2024）
                file_patterns = [
                    r'(\d{2,4}级[^\d_]+\d*班)',  # 如 "2022级计算机1班"
                    r'([a-zA-Z\u4e00-\u9fa5]+\d{4})',  # 如 "软件工程2301"（中文或英文+4位数字）
                    r'_([^\d_][^_]+)$',  # 如 "_计算机1班"
                    r'^([^\d_]+\d+班)',  # 如 "计算机1班"
                ]
                for pattern in file_patterns:
                    file_match = re.search(pattern, filename_no_ext)
                    if file_match:
                        candidate = file_match.group(1).strip()
                        # 检查是否包含排除关键词
                        if not any(kw in candidate for kw in exclude_keywords):
                            class_name = candidate
                            break

            # 1.4 查找列头中的班级列（用于从每行数据提取）
            # 注意：这里不再记录全局 class_col_idx，而是在后面为每组数据找班级列
            header_class_cols = []  # 存储所有班级列的位置
            for i in range(min(50, len(df))):
                row_values = [str(df.iloc[i, j]).strip() if pd.notna(df.iloc[i, j]) else ''
                              for j in range(len(df.columns))]
                # 找到列头行
                if any('学号' in v for v in row_values) and any('姓名' in v for v in row_values):
                    for j, cell_value in enumerate(row_values):
                        if cell_value in ['班级', '行政班']:
                            header_class_cols.append(j)
                    break

            # 确定班级获取方式
            use_class_column = len(header_class_cols) > 0 and not class_name
            if not class_name and not header_class_cols:
                warnings.append(f"工作表「{sheet}」: 未找到行政班/班级信息，班级列将留空")
                class_name = ""  # 行政班信息可选，留空继续处理

            # ===== 2. 动态查找列头行（支持多组并排格式） =====
            header_row = None
            col_groups = []  # 存储多组列映射，每组是一个 col_mapping

            key_patterns = {
                'student_id': ['学号'],
                'name': ['姓名'],
                'final_score': ['期末成绩', '期末', '期末考试'],
                'regular_score': ['平时成绩', '平时', '平时分'],
                'total_score': ['总成绩', '总评成绩', '总分', '成绩', '总评']
            }

            for i in range(min(50, len(df))):
                row_values = [str(df.iloc[i, j]).strip() if pd.notna(df.iloc[i, j]) else ''
                              for j in range(len(df.columns))]

                if any('学号' in v for v in row_values) and any('姓名' in v for v in row_values):
                    header_row = i

                    # 找到所有"学号"列的位置
                    student_id_cols = [j for j, v in enumerate(row_values) if '学号' in v]

                    for sid_col in student_id_cols:
                        col_mapping = {'student_id': sid_col}

                        # 在学号列之后查找其他列（直到下一个学号列或行尾）
                        next_sid_col = len(row_values)
                        for next_col in student_id_cols:
                            if next_col > sid_col:
                                next_sid_col = next_col
                                break

                        # 在 [sid_col, next_sid_col) 范围内查找其他列
                        for j in range(sid_col, next_sid_col):
                            cell_value = row_values[j].strip()

                            if '姓名' in cell_value and 'name' not in col_mapping:
                                col_mapping['name'] = j

                            if any(p in cell_value for p in key_patterns['final_score']) and 'final_score' not in col_mapping:
                                col_mapping['final_score'] = j

                            if any(p in cell_value for p in key_patterns['regular_score']) and 'regular_score' not in col_mapping:
                                col_mapping['regular_score'] = j

                            if 'total_score' not in col_mapping:
                                # 优先精确匹配，避免"成绩"匹配到"平时成绩"等
                                if any(p in cell_value for p in ['总成绩', '总评成绩', '总分']):
                                    col_mapping['total_score'] = j
                                elif cell_value in ['成绩', '总评']:
                                    col_mapping['total_score'] = j

                            # 为每组数据查找对应的班级列
                            if cell_value in ['班级', '行政班'] and 'class_col' not in col_mapping:
                                col_mapping['class_col'] = j

                        # 检查这组是否有完整的必需列
                        required_cols = ['student_id', 'name', 'final_score', 'regular_score', 'total_score']
                        missing_cols = [col for col in required_cols if col not in col_mapping]
                        if not missing_cols:
                            col_groups.append(col_mapping)
                        elif sid_col == student_id_cols[0]:  # 只为第一组记录缺失信息
                            col_name_map = {
                                'student_id': '学号', 'name': '姓名',
                                'final_score': '期末成绩', 'regular_score': '平时成绩',
                                'total_score': '总成绩'
                            }
                            missing_names = [col_name_map[c] for c in missing_cols]
                            warnings.append(f"工作表「{sheet}」: 缺少列「{'、'.join(missing_names)}」，已跳过")

                    break

            if header_row is None:
                warnings.append(f"工作表「{sheet}」: 未找到包含'学号'和'姓名'的列头行，已跳过")
                continue

            if not col_groups:
                # 如果前面没有添加具体的缺失列警告，则添加通用警告
                has_missing_warning = any(f"工作表「{sheet}」: 缺少列" in w for w in warnings)
                if not has_missing_warning:
                    warnings.append(f"工作表「{sheet}」: 未找到完整的成绩列组合，已跳过")
                continue

            # 如果有多组，统计数量（不再作为警告）
            if len(col_groups) > 1:
                multi_col_sheets += 1

            # ===== 3. 提取学生数据（支持多组） =====
            data_start_row = header_row + 1

            for i in range(data_start_row, len(df)):
                row = df.iloc[i]

                # 从每组中提取学生数据
                for col_mapping in col_groups:
                    student_id = str(row[col_mapping['student_id']]) if pd.notna(row[col_mapping['student_id']]) else ''

                    if is_valid_student_id(student_id):
                        name = row[col_mapping['name']]

                        # 获取班级信息：
                        # 1. 优先从当前组的班级列获取
                        # 2. 其次使用统一班级名（从顶部/工作表名称提取）
                        if 'class_col' in col_mapping:
                            student_class = str(row[col_mapping['class_col']]).strip() if pd.notna(row[col_mapping['class_col']]) else ""
                        elif use_class_column and header_class_cols:
                            # 如果没有组内班级列，尝试使用最近的全局班级列
                            # 找到距离当前学号列最近的班级列
                            sid_col = col_mapping['student_id']
                            closest_class_col = min(header_class_cols, key=lambda x: abs(x - sid_col))
                            student_class = str(row[closest_class_col]).strip() if pd.notna(row[closest_class_col]) else ""
                        else:
                            student_class = class_name if class_name else ""

                        final_raw = row[col_mapping['final_score']]
                        regular_raw = row[col_mapping['regular_score']]
                        total_raw = row[col_mapping['total_score']]

                        special_status = None
                        special_keywords = ['缺考', '缓考', '作弊', '取消', '免修', '旷考']

                        for raw_val in [final_raw, regular_raw, total_raw]:
                            if pd.notna(raw_val):
                                raw_str = str(raw_val).strip()
                                for keyword in special_keywords:
                                    if keyword in raw_str:
                                        special_status = raw_str
                                        break
                            if special_status:
                                break

                        all_empty = is_empty(final_raw) and is_empty(regular_raw) and is_empty(total_raw)

                        if all_empty:
                            special_status = '成绩为空'

                        if special_status:
                            all_students.append({
                                'class': student_class,
                                'student_id': student_id,
                                'name': name,
                                'final_score': None,
                                'regular_score': None,
                                'total_score': None,
                                'status': special_status
                            })
                        else:
                            try:
                                final_score = float(final_raw)
                                regular_score = float(regular_raw)
                                total_score = float(total_raw)

                                all_students.append({
                                    'class': student_class,
                                    'student_id': student_id,
                                    'name': name,
                                    'final_score': final_score,
                                    'regular_score': regular_score,
                                    'total_score': total_score,
                                    'status': None
                                })
                            except (ValueError, TypeError):
                                all_students.append({
                                    'class': student_class,
                                    'student_id': student_id,
                                    'name': name,
                                    'final_score': None,
                                    'regular_score': None,
                                    'total_score': None,
                                    'status': '成绩异常'
                                })

            processed_sheets += 1
            progress = 5 + int(25 * processed_sheets / max(total_sheets, 1))
            self._report_progress(f"正在处理工作表 {sheet}...", progress)

        # 添加多列统计汇总信息（仅当有多列工作表时）
        if multi_col_sheets > 0:
            warnings.insert(0, f"📊 检测到 {multi_col_sheets} 个工作表包含多组并排学生数据，已全部正确处理")

        return all_students, warnings

    def sort_students(self, students: list[dict]) -> list[dict]:
        """按行政班分组，按学号升序排序"""
        return sorted(students, key=lambda x: (x['class'], x['student_id']))

    def get_class_statistics(self, students: list[dict]) -> dict[str, int]:
        """获取班级统计信息"""
        return dict(Counter([s['class'] for s in students]))

    def create_workbook(self, output_file: str, students: list[dict]):
        """从零创建工作簿，填入学生数据并生成输出文件"""
        self._report_progress("正在创建工作簿...", 35)

        config = self.config
        ratio_1 = config.ratio_1
        ratio_2 = config.ratio_2
        ratio_3 = config.ratio_3

        # 创建新工作簿
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        ws_calc = wb.create_sheet('课程目标达成度计算', 0)
        ws_stat = wb.create_sheet('达成度统计', 1)
        wb.remove(default_sheet)

        # 定义样式
        black_font = Font(color="000000")
        bold_font = Font(color="000000", bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 计算需要的行数
        num_students = len(students)
        data_start_row = 3
        data_end_row = data_start_row + num_students - 1
        avg_row = data_end_row + 1

        self._report_progress("正在设置标题行...", 40)
        self._setup_calc_sheet_headers(ws_calc, ratio_1, ratio_2, ratio_3,
                                       bold_font, black_font, center_alignment, thin_border)

        self._report_progress("正在填入学生数据...", 45)
        self._fill_student_data(ws_calc, students, data_start_row, data_end_row,
                                black_font, center_alignment, thin_border)

        self._report_progress("正在计算达成度...", 60)
        self._fill_achievement_data(ws_calc, students, data_start_row, data_end_row,
                                    black_font, center_alignment, thin_border)

        self._report_progress("正在计算平均值...", 70)
        self._fill_average_row(ws_calc, avg_row, data_start_row, data_end_row,
                               black_font, center_alignment, right_alignment, thin_border)

        self._setup_column_widths(ws_calc)

        self._report_progress("正在创建统计页...", 75)
        self._setup_statistics_sheet(ws_stat, data_start_row, data_end_row,
                                     black_font, bold_font, center_alignment, thin_border)

        self._report_progress("正在创建图表...", 85)
        self._create_charts(ws_calc, ws_stat, data_start_row, data_end_row)

        self._report_progress("正在保存文件...", 95)
        try:
            wb.save(output_file)
        except PermissionError:
            raise PermissionError(f"无法保存文件，可能被其他程序占用或目录无写入权限")
        except Exception as e:
            raise IOError(f"保存文件失败: {str(e)}")
        self._report_progress("处理完成！", 100)

    def _setup_calc_sheet_headers(self, ws_calc, ratio_1, ratio_2, ratio_3,
                                  bold_font, black_font, center_alignment, thin_border):
        """设置课程目标达成度计算工作表的标题行"""
        config = self.config

        # 第一行
        ws_calc.merge_cells('A1:B1')
        ws_calc.cell(1, 1).border = thin_border
        ws_calc.cell(1, 2).border = thin_border

        ws_calc.cell(1, 3).value = ratio_1
        ws_calc.cell(1, 3).font = black_font
        ws_calc.cell(1, 3).alignment = center_alignment
        ws_calc.cell(1, 3).border = thin_border

        ws_calc.cell(1, 4).value = ratio_2
        ws_calc.cell(1, 4).font = black_font
        ws_calc.cell(1, 4).alignment = center_alignment
        ws_calc.cell(1, 4).border = thin_border

        ws_calc.cell(1, 5).value = ratio_3
        ws_calc.cell(1, 5).font = black_font
        ws_calc.cell(1, 5).alignment = center_alignment
        ws_calc.cell(1, 5).border = thin_border

        ws_calc.merge_cells('F1:H1')
        ws_calc.cell(1, 6).value = '成绩'
        ws_calc.cell(1, 6).font = bold_font
        ws_calc.cell(1, 6).alignment = center_alignment
        ws_calc.cell(1, 6).border = thin_border

        ws_calc.merge_cells('I1:J1')
        ws_calc.cell(1, 9).border = thin_border
        ws_calc.cell(1, 10).border = thin_border

        ws_calc.merge_cells('K1:L1')
        ws_calc.cell(1, 11).value = '平时成绩'
        ws_calc.cell(1, 11).font = bold_font
        ws_calc.cell(1, 11).alignment = center_alignment
        ws_calc.cell(1, 11).border = thin_border

        ws_calc.merge_cells('M1:N1')
        ws_calc.cell(1, 13).value = config.regular_score_ratio
        ws_calc.cell(1, 13).font = black_font
        ws_calc.cell(1, 13).alignment = center_alignment
        ws_calc.cell(1, 13).border = thin_border

        ws_calc.merge_cells('O1:P1')
        ws_calc.cell(1, 15).value = '期末成绩'
        ws_calc.cell(1, 15).font = bold_font
        ws_calc.cell(1, 15).alignment = center_alignment
        ws_calc.cell(1, 15).border = thin_border

        ws_calc.merge_cells('Q1:R1')
        ws_calc.cell(1, 17).value = config.final_score_ratio
        ws_calc.cell(1, 17).font = black_font
        ws_calc.cell(1, 17).alignment = center_alignment
        ws_calc.cell(1, 17).border = thin_border

        ws_calc.merge_cells('S1:V1')
        ws_calc.cell(1, 19).value = '总成绩'
        ws_calc.cell(1, 19).font = bold_font
        ws_calc.cell(1, 19).alignment = center_alignment
        ws_calc.cell(1, 19).border = thin_border

        ws_calc.merge_cells('W1:Y1')
        ws_calc.cell(1, 23).value = '达成度'
        ws_calc.cell(1, 23).font = bold_font
        ws_calc.cell(1, 23).alignment = center_alignment
        ws_calc.cell(1, 23).border = thin_border

        ws_calc.merge_cells('Z1:AB1')
        ws_calc.cell(1, 26).value = '达成度平均值'
        ws_calc.cell(1, 26).font = bold_font
        ws_calc.cell(1, 26).alignment = center_alignment
        ws_calc.cell(1, 26).border = thin_border

        ws_calc.merge_cells('AC1:AE1')
        ws_calc.cell(1, 29).value = '达成度期望值'
        ws_calc.cell(1, 29).font = bold_font
        ws_calc.cell(1, 29).alignment = center_alignment
        ws_calc.cell(1, 29).border = thin_border

        ws_calc.cell(1, 32).value = '算术平均值'
        ws_calc.cell(1, 32).font = bold_font
        ws_calc.cell(1, 32).alignment = center_alignment
        ws_calc.cell(1, 32).border = thin_border

        ws_calc.merge_cells('AG1:AG2')
        ws_calc.cell(1, 33).value = '总达成度平均值'
        ws_calc.cell(1, 33).font = bold_font
        ws_calc.cell(1, 33).alignment = center_alignment
        ws_calc.cell(1, 33).border = thin_border
        ws_calc.cell(2, 33).border = thin_border

        # 第二行
        row2_headers = [
            ('A', '班级'), ('B', '学号'), ('C', '目标一'), ('D', '目标二'), ('E', '目标三'),
            ('F', '平时'), ('G', '期末'), ('H', '总分'),
            ('I', '序号'), ('J', '姓名'),
            ('K', '目标1'), ('L', '目标2'), ('M', '目标3'), ('N', '平时'),
            ('O', '目标1'), ('P', '目标2'), ('Q', '目标3'), ('R', '期末'),
            ('S', '目标1'), ('T', '目标2'), ('U', '目标3'), ('V', '总分'),
            ('W', '目标1'), ('X', '目标2'), ('Y', '目标3'),
            ('Z', '目标1'), ('AA', '目标2'), ('AB', '目标3'),
            ('AC', '目标1'), ('AD', '目标2'), ('AE', '目标3'),
            ('AF', '总达成度')
        ]

        for col_letter, header in row2_headers:
            col_idx = openpyxl.utils.column_index_from_string(col_letter)
            ws_calc.cell(2, col_idx).value = header
            ws_calc.cell(2, col_idx).font = bold_font
            ws_calc.cell(2, col_idx).alignment = center_alignment
            ws_calc.cell(2, col_idx).border = thin_border

    def _fill_student_data(self, ws_calc, students, data_start_row, _data_end_row,
                           black_font, center_alignment, thin_border):
        """填入学生基本数据"""
        for idx, student in enumerate(students):
            row = data_start_row + idx
            is_special = student.get('status') is not None

            # A列: 班级
            ws_calc.cell(row, 1).value = student['class']
            ws_calc.cell(row, 1).font = black_font
            ws_calc.cell(row, 1).alignment = center_alignment
            ws_calc.cell(row, 1).border = thin_border

            # B列: 学号
            ws_calc.cell(row, 2).value = student['student_id']
            ws_calc.cell(row, 2).font = black_font
            ws_calc.cell(row, 2).alignment = center_alignment
            ws_calc.cell(row, 2).border = thin_border

            # I列: 序号
            ws_calc.cell(row, 9).value = idx + 1
            ws_calc.cell(row, 9).font = black_font
            ws_calc.cell(row, 9).alignment = center_alignment
            ws_calc.cell(row, 9).border = thin_border

            # J列: 姓名
            ws_calc.cell(row, 10).value = student['name']
            ws_calc.cell(row, 10).font = black_font
            ws_calc.cell(row, 10).alignment = center_alignment
            ws_calc.cell(row, 10).border = thin_border

            if is_special:
                for col in range(3, 6):
                    ws_calc.cell(row, col).border = thin_border
                ws_calc.cell(row, 6).border = thin_border
                ws_calc.cell(row, 7).border = thin_border
                ws_calc.cell(row, 8).value = student['status']
                ws_calc.cell(row, 8).font = black_font
                ws_calc.cell(row, 8).alignment = center_alignment
                ws_calc.cell(row, 8).border = thin_border
                for col in range(11, 26):
                    ws_calc.cell(row, col).border = thin_border
            else:
                # C列: 目标一
                ws_calc.cell(row, 3).value = f'=ROUND(H{row}*$C$1/100,0)'
                ws_calc.cell(row, 3).font = black_font
                ws_calc.cell(row, 3).alignment = center_alignment
                ws_calc.cell(row, 3).border = thin_border

                # D列: 目标二
                ws_calc.cell(row, 4).value = f'=ROUND(H{row}*$D$1/100,0)'
                ws_calc.cell(row, 4).font = black_font
                ws_calc.cell(row, 4).alignment = center_alignment
                ws_calc.cell(row, 4).border = thin_border

                # E列: 目标三
                ws_calc.cell(row, 5).value = f'=ROUND(H{row}*$E$1/100,0)'
                ws_calc.cell(row, 5).font = black_font
                ws_calc.cell(row, 5).alignment = center_alignment
                ws_calc.cell(row, 5).border = thin_border

                # F列: 平时成绩
                ws_calc.cell(row, 6).value = student['regular_score']
                ws_calc.cell(row, 6).font = black_font
                ws_calc.cell(row, 6).alignment = center_alignment
                ws_calc.cell(row, 6).border = thin_border
                ws_calc.cell(row, 6).number_format = '0.00'

                # G列: 期末成绩
                ws_calc.cell(row, 7).value = student['final_score']
                ws_calc.cell(row, 7).font = black_font
                ws_calc.cell(row, 7).alignment = center_alignment
                ws_calc.cell(row, 7).border = thin_border
                ws_calc.cell(row, 7).number_format = '0.00'

                # H列: 总成绩
                ws_calc.cell(row, 8).value = student['total_score']
                ws_calc.cell(row, 8).font = black_font
                ws_calc.cell(row, 8).alignment = center_alignment
                ws_calc.cell(row, 8).border = thin_border
                ws_calc.cell(row, 8).number_format = '0.00'

                # K-V列: 达成率计算
                formulas = [
                    (11, f'=(ROUND(F{row}*$C$1/100,0)/$C$1)*100'),
                    (12, f'=(ROUND(F{row}*$D$1/100,0)/$D$1)*100'),
                    (13, f'=(ROUND(F{row}*$E$1/100,0)/$E$1)*100'),
                    (14, f'=F{row}'),
                    (15, f'=(ROUND(G{row}*$C$1/100,0)/$C$1)*100'),
                    (16, f'=(ROUND(G{row}*$D$1/100,0)/$D$1)*100'),
                    (17, f'=(ROUND(G{row}*$E$1/100,0)/$E$1)*100'),
                    (18, f'=G{row}'),
                    (19, f'=K{row}*$M$1/100+O{row}*$Q$1/100'),
                    (20, f'=L{row}*$M$1/100+P{row}*$Q$1/100'),
                    (21, f'=M{row}*$M$1/100+Q{row}*$Q$1/100'),
                    (22, f'=H{row}'),
                    (23, f'=S{row}/100'),
                    (24, f'=T{row}/100'),
                    (25, f'=U{row}/100'),
                ]

                for col, formula in formulas:
                    ws_calc.cell(row, col).value = formula
                    ws_calc.cell(row, col).font = black_font
                    ws_calc.cell(row, col).alignment = center_alignment
                    ws_calc.cell(row, col).border = thin_border
                    ws_calc.cell(row, col).number_format = '0.00'

    def _fill_achievement_data(self, ws_calc, students, data_start_row, data_end_row,
                               black_font, center_alignment, thin_border):
        """填入达成度数据（Z-AG列）"""
        config = self.config

        for idx, student in enumerate(students):
            row = data_start_row + idx
            is_special = student.get('status') is not None

            # AC-AE列: 达成度期望值
            for col in [29, 30, 31]:
                ws_calc.cell(row, col).value = config.achievement_expectation
                ws_calc.cell(row, col).font = black_font
                ws_calc.cell(row, col).alignment = center_alignment
                ws_calc.cell(row, col).border = thin_border
                ws_calc.cell(row, col).number_format = '0.00'

            if is_special:
                for col in [26, 27, 28, 32, 33]:
                    ws_calc.cell(row, col).border = thin_border
            else:
                # Z列: 目标1达成度平均值
                ws_calc.cell(row, 26).value = f'=AVERAGE(W${data_start_row}:W${data_end_row})'
                ws_calc.cell(row, 26).font = black_font
                ws_calc.cell(row, 26).alignment = center_alignment
                ws_calc.cell(row, 26).border = thin_border
                ws_calc.cell(row, 26).number_format = '0.00'

                # AA列
                ws_calc.cell(row, 27).value = f'=AVERAGE(X${data_start_row}:X${data_end_row})'
                ws_calc.cell(row, 27).font = black_font
                ws_calc.cell(row, 27).alignment = center_alignment
                ws_calc.cell(row, 27).border = thin_border
                ws_calc.cell(row, 27).number_format = '0.00'

                # AB列
                ws_calc.cell(row, 28).value = f'=AVERAGE(Y${data_start_row}:Y${data_end_row})'
                ws_calc.cell(row, 28).font = black_font
                ws_calc.cell(row, 28).alignment = center_alignment
                ws_calc.cell(row, 28).border = thin_border
                ws_calc.cell(row, 28).number_format = '0.00'

                # AF列: 总达成度
                ws_calc.cell(row, 32).value = f'=V{row}/100'
                ws_calc.cell(row, 32).font = black_font
                ws_calc.cell(row, 32).alignment = center_alignment
                ws_calc.cell(row, 32).border = thin_border
                ws_calc.cell(row, 32).number_format = '0.00'

                # AG列: 总达成度平均值
                ws_calc.cell(row, 33).value = f'=AVERAGE(AF${data_start_row}:AF${data_end_row})'
                ws_calc.cell(row, 33).font = black_font
                ws_calc.cell(row, 33).alignment = center_alignment
                ws_calc.cell(row, 33).border = thin_border
                ws_calc.cell(row, 33).number_format = '0.00'

    def _fill_average_row(self, ws_calc, avg_row, data_start_row, data_end_row,
                          black_font, center_alignment, right_alignment, thin_border):
        """填入平均值行"""
        ws_calc.merge_cells(f'A{avg_row}:B{avg_row}')
        ws_calc.cell(avg_row, 1).value = '（平均值）'
        ws_calc.cell(avg_row, 1).font = black_font
        ws_calc.cell(avg_row, 1).alignment = center_alignment
        ws_calc.cell(avg_row, 1).border = thin_border
        ws_calc.cell(avg_row, 2).border = thin_border

        # C-H列
        for col in range(3, 9):
            col_letter = get_column_letter(col)
            ws_calc.cell(avg_row, col).value = f'=AVERAGE({col_letter}{data_start_row}:{col_letter}{data_end_row})'
            ws_calc.cell(avg_row, col).font = black_font
            ws_calc.cell(avg_row, col).alignment = right_alignment
            ws_calc.cell(avg_row, col).border = thin_border
            ws_calc.cell(avg_row, col).number_format = '0.00'

        # K-V列
        for col in range(11, 23):
            col_letter = get_column_letter(col)
            ws_calc.cell(avg_row, col).value = f'=AVERAGE({col_letter}{data_start_row}:{col_letter}{data_end_row})'
            ws_calc.cell(avg_row, col).font = black_font
            ws_calc.cell(avg_row, col).alignment = right_alignment
            ws_calc.cell(avg_row, col).border = thin_border
            ws_calc.cell(avg_row, col).number_format = '0.00'

        # W-Y列
        for col in range(23, 26):
            col_letter = get_column_letter(col)
            ws_calc.cell(avg_row, col).value = f'=AVERAGE({col_letter}{data_start_row}:{col_letter}{data_end_row})'
            ws_calc.cell(avg_row, col).font = black_font
            ws_calc.cell(avg_row, col).alignment = right_alignment
            ws_calc.cell(avg_row, col).border = thin_border
            ws_calc.cell(avg_row, col).number_format = '0.00'

        # AF列
        ws_calc.cell(avg_row, 32).value = f'=AVERAGE(AF{data_start_row}:AF{data_end_row})'
        ws_calc.cell(avg_row, 32).font = black_font
        ws_calc.cell(avg_row, 32).alignment = right_alignment
        ws_calc.cell(avg_row, 32).border = thin_border
        ws_calc.cell(avg_row, 32).number_format = '0.00'

    def _setup_column_widths(self, ws_calc):
        """设置列宽"""
        numeric_width = 11

        for col in range(3, 9):
            ws_calc.column_dimensions[get_column_letter(col)].width = numeric_width
        for col in range(11, 26):
            ws_calc.column_dimensions[get_column_letter(col)].width = numeric_width
        for col in range(26, 29):
            ws_calc.column_dimensions[get_column_letter(col)].width = numeric_width
        for col in range(29, 32):
            ws_calc.column_dimensions[get_column_letter(col)].width = numeric_width

        ws_calc.column_dimensions['AF'].width = numeric_width + 2
        ws_calc.column_dimensions['AG'].width = 16.5
        ws_calc.column_dimensions['A'].width = 13
        ws_calc.column_dimensions['B'].width = 13
        ws_calc.column_dimensions['J'].width = 9
        ws_calc.column_dimensions['I'].width = 6

    def _setup_statistics_sheet(self, ws_stat, data_start_row, data_end_row,
                                black_font, bold_font, center_alignment, thin_border):
        """设置达成度统计工作表"""
        # 第一行标题
        ws_stat.cell(1, 1).value = '达成度'
        ws_stat.cell(1, 1).font = bold_font
        ws_stat.cell(1, 1).alignment = center_alignment
        ws_stat.cell(1, 1).border = thin_border

        ws_stat.cell(1, 2).value = '达成情况'
        ws_stat.cell(1, 2).font = bold_font
        ws_stat.cell(1, 2).alignment = center_alignment
        ws_stat.cell(1, 2).border = thin_border

        ws_stat.merge_cells('C1:D1')
        ws_stat.cell(1, 3).value = '目标1'
        ws_stat.cell(1, 3).font = bold_font
        ws_stat.cell(1, 3).alignment = center_alignment
        ws_stat.cell(1, 3).border = thin_border
        ws_stat.cell(1, 4).border = thin_border

        ws_stat.merge_cells('E1:F1')
        ws_stat.cell(1, 5).value = '目标2'
        ws_stat.cell(1, 5).font = bold_font
        ws_stat.cell(1, 5).alignment = center_alignment
        ws_stat.cell(1, 5).border = thin_border
        ws_stat.cell(1, 6).border = thin_border

        ws_stat.merge_cells('G1:H1')
        ws_stat.cell(1, 7).value = '目标3'
        ws_stat.cell(1, 7).font = bold_font
        ws_stat.cell(1, 7).alignment = center_alignment
        ws_stat.cell(1, 7).border = thin_border
        ws_stat.cell(1, 8).border = thin_border

        # 第二行
        ws_stat.cell(2, 1).border = thin_border
        ws_stat.cell(2, 2).border = thin_border
        for col, header in [(3, '人数'), (4, '占比'), (5, '人数'), (6, '占比'), (7, '人数'), (8, '占比')]:
            ws_stat.cell(2, col).value = header
            ws_stat.cell(2, col).font = bold_font
            ws_stat.cell(2, col).alignment = center_alignment
            ws_stat.cell(2, col).border = thin_border

        # 达成度标准行
        standards = [
            (3, '>0.8', '完全达成'),
            (4, '0.6-0.8', '较好达成'),
            (5, '0.5-0.6', '基本达成'),
            (6, '0.4-0.5', '较少达成'),
            (7, '<0.4', '没有达成')
        ]

        for row, level, desc in standards:
            ws_stat.cell(row, 1).value = level
            ws_stat.cell(row, 1).font = black_font
            ws_stat.cell(row, 1).alignment = center_alignment
            ws_stat.cell(row, 1).border = thin_border

            ws_stat.cell(row, 2).value = desc
            ws_stat.cell(row, 2).font = black_font
            ws_stat.cell(row, 2).alignment = center_alignment
            ws_stat.cell(row, 2).border = thin_border

        # 人数统计公式
        ws_stat.cell(3, 3).value = f'=COUNTIF(\'课程目标达成度计算\'!W{data_start_row}:W{data_end_row},">0.8")'
        ws_stat.cell(4, 3).value = f'=COUNTIFS(\'课程目标达成度计算\'!W{data_start_row}:W{data_end_row},">=0.6",\'课程目标达成度计算\'!W{data_start_row}:W{data_end_row},"<=0.8")'
        ws_stat.cell(5, 3).value = f'=COUNTIFS(\'课程目标达成度计算\'!W{data_start_row}:W{data_end_row},">=0.5",\'课程目标达成度计算\'!W{data_start_row}:W{data_end_row},"<0.6")'
        ws_stat.cell(6, 3).value = f'=COUNTIFS(\'课程目标达成度计算\'!W{data_start_row}:W{data_end_row},">=0.4",\'课程目标达成度计算\'!W{data_start_row}:W{data_end_row},"<0.5")'
        ws_stat.cell(7, 3).value = f'=COUNTIF(\'课程目标达成度计算\'!W{data_start_row}:W{data_end_row},"<0.4")'

        ws_stat.cell(3, 5).value = f'=COUNTIF(\'课程目标达成度计算\'!X{data_start_row}:X{data_end_row},">0.8")'
        ws_stat.cell(4, 5).value = f'=COUNTIFS(\'课程目标达成度计算\'!X{data_start_row}:X{data_end_row},">=0.6",\'课程目标达成度计算\'!X{data_start_row}:X{data_end_row},"<=0.8")'
        ws_stat.cell(5, 5).value = f'=COUNTIFS(\'课程目标达成度计算\'!X{data_start_row}:X{data_end_row},">=0.5",\'课程目标达成度计算\'!X{data_start_row}:X{data_end_row},"<0.6")'
        ws_stat.cell(6, 5).value = f'=COUNTIFS(\'课程目标达成度计算\'!X{data_start_row}:X{data_end_row},">=0.4",\'课程目标达成度计算\'!X{data_start_row}:X{data_end_row},"<0.5")'
        ws_stat.cell(7, 5).value = f'=COUNTIF(\'课程目标达成度计算\'!X{data_start_row}:X{data_end_row},"<0.4")'

        ws_stat.cell(3, 7).value = f'=COUNTIF(\'课程目标达成度计算\'!Y{data_start_row}:Y{data_end_row},">0.8")'
        ws_stat.cell(4, 7).value = f'=COUNTIFS(\'课程目标达成度计算\'!Y{data_start_row}:Y{data_end_row},">=0.6",\'课程目标达成度计算\'!Y{data_start_row}:Y{data_end_row},"<=0.8")'
        ws_stat.cell(5, 7).value = f'=COUNTIFS(\'课程目标达成度计算\'!Y{data_start_row}:Y{data_end_row},">=0.5",\'课程目标达成度计算\'!Y{data_start_row}:Y{data_end_row},"<0.6")'
        ws_stat.cell(6, 7).value = f'=COUNTIFS(\'课程目标达成度计算\'!Y{data_start_row}:Y{data_end_row},">=0.4",\'课程目标达成度计算\'!Y{data_start_row}:Y{data_end_row},"<0.5")'
        ws_stat.cell(7, 7).value = f'=COUNTIF(\'课程目标达成度计算\'!Y{data_start_row}:Y{data_end_row},"<0.4")'

        # 占比公式和样式
        for row in range(3, 8):
            for col in [3, 5, 7]:
                ws_stat.cell(row, col).font = black_font
                ws_stat.cell(row, col).alignment = center_alignment
                ws_stat.cell(row, col).border = thin_border

            ws_stat.cell(row, 4).value = f'=C{row}/COUNT(\'课程目标达成度计算\'!W${data_start_row}:W${data_end_row})'
            ws_stat.cell(row, 4).font = black_font
            ws_stat.cell(row, 4).alignment = center_alignment
            ws_stat.cell(row, 4).border = thin_border
            ws_stat.cell(row, 4).number_format = '0.00%'

            ws_stat.cell(row, 6).value = f'=E{row}/COUNT(\'课程目标达成度计算\'!X${data_start_row}:X${data_end_row})'
            ws_stat.cell(row, 6).font = black_font
            ws_stat.cell(row, 6).alignment = center_alignment
            ws_stat.cell(row, 6).border = thin_border
            ws_stat.cell(row, 6).number_format = '0.00%'

            ws_stat.cell(row, 8).value = f'=G{row}/COUNT(\'课程目标达成度计算\'!Y${data_start_row}:Y${data_end_row})'
            ws_stat.cell(row, 8).font = black_font
            ws_stat.cell(row, 8).alignment = center_alignment
            ws_stat.cell(row, 8).border = thin_border
            ws_stat.cell(row, 8).number_format = '0.00%'

        # 设置列宽
        ws_stat.column_dimensions['A'].width = 11
        ws_stat.column_dimensions['B'].width = 11
        for col in ['C', 'D', 'E', 'F', 'G', 'H']:
            ws_stat.column_dimensions[col].width = 7

    def _create_charts(self, ws_calc, ws_stat, data_start_row, data_end_row):
        """创建所有图表"""
        # 根据学生数量动态计算X轴标签间隔
        num_students = data_end_row - data_start_row + 1
        tick_skip = max(1, num_students // 10)  # 大约保持显示10个标签

        # 课程目标达成度计算页的折线图
        chart_configs = [
            {'title': '目标1达成度', 'y_col': 23, 'avg_col': 26, 'exp_col': 29},
            {'title': '目标2达成度', 'y_col': 24, 'avg_col': 27, 'exp_col': 30},
            {'title': '目标3达成度', 'y_col': 25, 'avg_col': 28, 'exp_col': 31},
            {'title': '总达成度', 'y_col': 32, 'avg_col': 33, 'exp_col': 29},
        ]

        chart_width = 18
        chart_height = 12
        start_col = 36
        col_gap = 12
        row_gap = 24
        row1_start = 2

        for i, config in enumerate(chart_configs):
            chart = LineChart()
            chart.title = config['title']
            chart.style = 10
            chart.x_axis.title = '学生序号'
            chart.y_axis.title = '达成度'
            chart.legend = None

            chart.y_axis.scaling.min = 0
            chart.y_axis.scaling.max = 1

            gridline_props = GraphicalProperties(
                ln=LineProperties(solidFill='C0C0C0', w=9525)
            )
            chart.x_axis.majorGridlines = ChartLines(spPr=gridline_props)
            chart.y_axis.majorGridlines = ChartLines(spPr=gridline_props)

            chart.x_axis.tickLblSkip = tick_skip
            chart.x_axis.tickMarkSkip = tick_skip

            x_values = Reference(ws_calc, min_col=9, min_row=data_start_row, max_row=data_end_row)

            y_values = Reference(ws_calc, min_col=config['y_col'], min_row=data_start_row - 1, max_row=data_end_row)
            chart.add_data(y_values, titles_from_data=True)

            avg_values = Reference(ws_calc, min_col=config['avg_col'], min_row=data_start_row - 1, max_row=data_end_row)
            chart.add_data(avg_values, titles_from_data=True)

            exp_values = Reference(ws_calc, min_col=config['exp_col'], min_row=data_start_row - 1, max_row=data_end_row)
            chart.add_data(exp_values, titles_from_data=True)

            chart.set_categories(x_values)

            if len(chart.series) >= 1:
                chart.series[0].marker = Marker(symbol='circle', size=5)
                chart.series[0].graphicalProperties.line = LineProperties(noFill=True)

            if len(chart.series) >= 2:
                chart.series[1].marker = Marker(symbol='none')
                chart.series[1].graphicalProperties.line = LineProperties(
                    solidFill='00FF00', w=25000, cmpd='dbl', prstDash='sysDot'
                )

            if len(chart.series) >= 3:
                chart.series[2].marker = Marker(symbol='none')
                chart.series[2].graphicalProperties.line = LineProperties(
                    solidFill='FF0000', w=25000, cmpd='dbl', prstDash='sysDot'
                )

            col_offset = (i % 2) * col_gap
            row_offset = (i // 2) * row_gap
            chart.anchor = f'{get_column_letter(start_col + col_offset)}{row1_start + row_offset}'
            chart.width = chart_width
            chart.height = chart_height

            ws_calc.add_chart(chart)

        # 达成度统计页的柱状图
        stat_chart_configs = [
            {'title': '目标1达成度人数占比统计', 'data_col': 4, 'anchor_col': 1},
            {'title': '目标2达成度人数占比统计', 'data_col': 6, 'anchor_col': 9},
            {'title': '目标3达成度人数占比统计', 'data_col': 8, 'anchor_col': 16},
        ]

        stat_chart_width = 10
        stat_chart_height = 10
        stat_start_row = 9

        for i, config in enumerate(stat_chart_configs):
            chart = BarChart()
            chart.title = config['title']
            chart.style = 10
            chart.type = 'col'
            chart.grouping = 'clustered'
            chart.legend = None

            chart.y_axis.scaling.min = 0
            chart.y_axis.scaling.max = 1
            chart.y_axis.numFmt = '0%'

            data = Reference(ws_stat, min_col=config['data_col'], min_row=2, max_row=7)
            cats = Reference(ws_stat, min_col=2, min_row=3, max_row=7)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            chart.x_axis.txPr = RichText(
                bodyPr=RichTextProperties(rot=0),
                p=[Paragraph(
                    pPr=ParagraphProperties(
                        defRPr=CharacterProperties(sz=900)
                    )
                )]
            )

            chart.anchor = f'{get_column_letter(config["anchor_col"])}{stat_start_row}'
            chart.width = stat_chart_width
            chart.height = stat_chart_height

            ws_stat.add_chart(chart)

    def process_file(self, input_file: str, output_file: str) -> dict:
        """处理单个文件

        Args:
            input_file: 输入的成绩单文件路径
            output_file: 输出的达成度报告文件路径

        Returns:
            处理结果信息，包含 total_students, class_statistics, output_file, warnings
        """
        # 1. 提取学生数据
        students, warnings = self.extract_students_from_grades(input_file)

        if not students:
            # 构建详细的错误信息
            error_msg = "未能从文件中提取到任何学生数据"
            if warnings:
                error_msg += "\n原因:\n" + "\n".join(f"  - {w}" for w in warnings)
            raise ValueError(error_msg)

        # 2. 排序
        self._report_progress("正在排序学生数据...", 30)
        students = self.sort_students(students)

        # 3. 获取统计信息
        class_stats = self.get_class_statistics(students)

        # 4. 创建工作簿
        self.create_workbook(output_file, students)

        return {
            'total_students': len(students),
            'class_statistics': class_stats,
            'output_file': output_file,
            'warnings': warnings  # 返回警告信息
        }
