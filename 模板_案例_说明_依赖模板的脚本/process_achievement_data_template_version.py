#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
达成度数据处理脚本
将成绩数据导入达成度数据模板，进行统计计算和绘图
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.chart import BarChart, ScatterChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.axis import Scaling
from openpyxl.utils import get_column_letter
from copy import copy
import re


def extract_students_from_grades(grades_file):
    """从成绩文件中提取所有学生数据"""
    xl = pd.ExcelFile(grades_file)
    all_students = []

    for sheet in xl.sheet_names:
        if sheet == 'Sheet1':
            continue

        df = pd.read_excel(xl, sheet_name=sheet, header=None)

        # 提取行政班名称
        class_info = str(df.iloc[2, 0]) if len(df) > 2 else ''
        if '行政班：' in class_info:
            class_name = class_info.split('行政班：')[1].split('(')[0].strip()
        else:
            continue

        # 找到数据行(列头在第4行索引4，数据从第5行索引5开始)
        for i in range(5, len(df)):
            row = df.iloc[i]
            student_id = str(row[1]) if pd.notna(row[1]) else ''

            # 检查是否是有效学生数据行（学号为纯数字且长度大于8）
            if student_id.isdigit() and len(student_id) > 8:
                # 处理可能的非数字成绩（如"缓考"）
                final_score = row[3]
                regular_score = row[4]
                total_score = row[5]

                # 跳过缓考或其他无效成绩的学生
                try:
                    final_score = float(final_score)
                    regular_score = float(regular_score)
                    total_score = float(total_score)
                except (ValueError, TypeError):
                    continue

                all_students.append({
                    'class': class_name,
                    'student_id': student_id,
                    'name': row[2],
                    'final_score': final_score,
                    'regular_score': regular_score,
                    'total_score': total_score
                })

    return all_students


def sort_students(students):
    """按行政班分组，按学号升序排序"""
    # 先按班级排序，再按学号排序
    return sorted(students, key=lambda x: (x['class'], x['student_id']))


def process_template(template_file, output_file, students):
    """处理模板文件，填入学生数据并生成输出文件"""

    # 加载模板
    wb = openpyxl.load_workbook(template_file)
    ws_calc = wb['课程目标达成度计算']
    ws_stat = wb['达成度统计']

    # 获取达成度占比 (C1, D1, E1)
    ratio_1 = ws_calc.cell(1, 3).value  # 目标一占比 (50)
    ratio_2 = ws_calc.cell(1, 4).value  # 目标二占比 (30)
    ratio_3 = ws_calc.cell(1, 5).value  # 目标三占比 (20)

    print(f"达成度占比: 目标一={ratio_1}%, 目标二={ratio_2}%, 目标三={ratio_3}%")

    # 定义样式
    black_font = Font(color="000000")
    bold_font = Font(color="000000", bold=True)  # 加粗字体
    center_alignment = Alignment(horizontal='center', vertical='center')
    right_alignment = Alignment(horizontal='right', vertical='center')  # 右对齐
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 计算需要的行数
    num_students = len(students)
    data_start_row = 3
    data_end_row = data_start_row + num_students - 1
    avg_row = data_end_row + 1  # 平均值行

    print(f"学生数量: {num_students}")
    print(f"数据行: {data_start_row} - {data_end_row}")
    print(f"平均值行: {avg_row}")

    # 定义需要保留两位小数的列（O-Y列, AA-AG列）
    decimal_cols = list(range(15, 26)) + list(range(27, 34))  # O-Y (15-25), AA-AG (27-33)

    # 填入学生数据
    for idx, student in enumerate(students):
        row = data_start_row + idx

        # A列: 班级
        ws_calc.cell(row, 1).value = student['class']
        ws_calc.cell(row, 1).font = black_font
        ws_calc.cell(row, 1).alignment = center_alignment
        ws_calc.cell(row, 1).border = thin_border

        # B列: 学号
        ws_calc.cell(row, 2).value = student['student_id']
        ws_calc.cell(row, 2).font = black_font
        ws_calc.cell(row, 2).alignment = center_alignment
        ws_calc.cell(row, 2).border = thin_border

        # C列: 目标一 = ROUND(总成绩 * $C$1 / 100, 0) - 基于总成绩计算
        ws_calc.cell(row, 3).value = f'=ROUND(H{row}*$C$1/100,0)'
        ws_calc.cell(row, 3).font = black_font
        ws_calc.cell(row, 3).alignment = center_alignment
        ws_calc.cell(row, 3).border = thin_border

        # D列: 目标二 = ROUND(总成绩 * $D$1 / 100, 0) - 基于总成绩计算
        ws_calc.cell(row, 4).value = f'=ROUND(H{row}*$D$1/100,0)'
        ws_calc.cell(row, 4).font = black_font
        ws_calc.cell(row, 4).alignment = center_alignment
        ws_calc.cell(row, 4).border = thin_border

        # E列: 目标三 = ROUND(总成绩 * $E$1 / 100, 0) - 基于总成绩计算
        ws_calc.cell(row, 5).value = f'=ROUND(H{row}*$E$1/100,0)'
        ws_calc.cell(row, 5).font = black_font
        ws_calc.cell(row, 5).alignment = center_alignment
        ws_calc.cell(row, 5).border = thin_border

        # F列: 平时成绩
        ws_calc.cell(row, 6).value = student['regular_score']
        ws_calc.cell(row, 6).font = black_font
        ws_calc.cell(row, 6).alignment = center_alignment
        ws_calc.cell(row, 6).border = thin_border
        ws_calc.cell(row, 6).number_format = '0.00'

        # G列: 期末成绩
        ws_calc.cell(row, 7).value = student['final_score']
        ws_calc.cell(row, 7).font = black_font
        ws_calc.cell(row, 7).alignment = center_alignment
        ws_calc.cell(row, 7).border = thin_border
        ws_calc.cell(row, 7).number_format = '0.00'

        # H列: 总成绩
        ws_calc.cell(row, 8).value = student['total_score']
        ws_calc.cell(row, 8).font = black_font
        ws_calc.cell(row, 8).alignment = center_alignment
        ws_calc.cell(row, 8).border = thin_border
        ws_calc.cell(row, 8).number_format = '0.00'

        # I列: 序号
        ws_calc.cell(row, 9).value = idx + 1
        ws_calc.cell(row, 9).font = black_font
        ws_calc.cell(row, 9).alignment = center_alignment
        ws_calc.cell(row, 9).border = thin_border

        # J列: 姓名
        ws_calc.cell(row, 10).value = student['name']
        ws_calc.cell(row, 10).font = black_font
        ws_calc.cell(row, 10).alignment = center_alignment
        ws_calc.cell(row, 10).border = thin_border

        # K列: 平时成绩目标1达成率 = (ROUND(F*C1/100,0)/C1)*100
        ws_calc.cell(row, 11).value = f'=(ROUND(F{row}*$C$1/100,0)/$C$1)*100'
        ws_calc.cell(row, 11).font = black_font
        ws_calc.cell(row, 11).alignment = center_alignment
        ws_calc.cell(row, 11).border = thin_border
        ws_calc.cell(row, 11).number_format = '0.00'

        # L列: 平时成绩目标2达成率 = (ROUND(F*D1/100,0)/D1)*100
        ws_calc.cell(row, 12).value = f'=(ROUND(F{row}*$D$1/100,0)/$D$1)*100'
        ws_calc.cell(row, 12).font = black_font
        ws_calc.cell(row, 12).alignment = center_alignment
        ws_calc.cell(row, 12).border = thin_border
        ws_calc.cell(row, 12).number_format = '0.00'

        # M列: 平时成绩目标3达成率 = (ROUND(F*E1/100,0)/E1)*100
        ws_calc.cell(row, 13).value = f'=(ROUND(F{row}*$E$1/100,0)/$E$1)*100'
        ws_calc.cell(row, 13).font = black_font
        ws_calc.cell(row, 13).alignment = center_alignment
        ws_calc.cell(row, 13).border = thin_border
        ws_calc.cell(row, 13).number_format = '0.00'

        # N列: 平时成绩 = F列原值
        ws_calc.cell(row, 14).value = f'=F{row}'
        ws_calc.cell(row, 14).font = black_font
        ws_calc.cell(row, 14).alignment = center_alignment
        ws_calc.cell(row, 14).border = thin_border
        ws_calc.cell(row, 14).number_format = '0.00'

        # O列: 期末成绩目标1达成率 = (ROUND(G*C1/100,0)/C1)*100
        ws_calc.cell(row, 15).value = f'=(ROUND(G{row}*$C$1/100,0)/$C$1)*100'
        ws_calc.cell(row, 15).font = black_font
        ws_calc.cell(row, 15).alignment = center_alignment
        ws_calc.cell(row, 15).border = thin_border
        ws_calc.cell(row, 15).number_format = '0.00'

        # P列: 期末成绩目标2达成率 = (ROUND(G*D1/100,0)/D1)*100
        ws_calc.cell(row, 16).value = f'=(ROUND(G{row}*$D$1/100,0)/$D$1)*100'
        ws_calc.cell(row, 16).font = black_font
        ws_calc.cell(row, 16).alignment = center_alignment
        ws_calc.cell(row, 16).border = thin_border
        ws_calc.cell(row, 16).number_format = '0.00'

        # Q列: 期末成绩目标3达成率 = (ROUND(G*E1/100,0)/E1)*100
        ws_calc.cell(row, 17).value = f'=(ROUND(G{row}*$E$1/100,0)/$E$1)*100'
        ws_calc.cell(row, 17).font = black_font
        ws_calc.cell(row, 17).alignment = center_alignment
        ws_calc.cell(row, 17).border = thin_border
        ws_calc.cell(row, 17).number_format = '0.00'

        # R列: 期末成绩 = G列原值
        ws_calc.cell(row, 18).value = f'=G{row}'
        ws_calc.cell(row, 18).font = black_font
        ws_calc.cell(row, 18).alignment = center_alignment
        ws_calc.cell(row, 18).border = thin_border
        ws_calc.cell(row, 18).number_format = '0.00'

        # S列: 总成绩目标1 = K*平时比例+O*期末比例 (动态引用$M$1和$Q$1)
        ws_calc.cell(row, 19).value = f'=K{row}*$M$1/100+O{row}*$Q$1/100'
        ws_calc.cell(row, 19).font = black_font
        ws_calc.cell(row, 19).alignment = center_alignment
        ws_calc.cell(row, 19).border = thin_border
        ws_calc.cell(row, 19).number_format = '0.00'

        # T列: 总成绩目标2 = L*平时比例+P*期末比例 (动态引用$M$1和$Q$1)
        ws_calc.cell(row, 20).value = f'=L{row}*$M$1/100+P{row}*$Q$1/100'
        ws_calc.cell(row, 20).font = black_font
        ws_calc.cell(row, 20).alignment = center_alignment
        ws_calc.cell(row, 20).border = thin_border
        ws_calc.cell(row, 20).number_format = '0.00'

        # U列: 总成绩目标3 = M*平时比例+Q*期末比例 (动态引用$M$1和$Q$1)
        ws_calc.cell(row, 21).value = f'=M{row}*$M$1/100+Q{row}*$Q$1/100'
        ws_calc.cell(row, 21).font = black_font
        ws_calc.cell(row, 21).alignment = center_alignment
        ws_calc.cell(row, 21).border = thin_border
        ws_calc.cell(row, 21).number_format = '0.00'

        # V列: 总成绩 = H列
        ws_calc.cell(row, 22).value = f'=H{row}'
        ws_calc.cell(row, 22).font = black_font
        ws_calc.cell(row, 22).alignment = center_alignment
        ws_calc.cell(row, 22).border = thin_border
        ws_calc.cell(row, 22).number_format = '0.00'

        # W列: 达成度目标1 = S/100
        ws_calc.cell(row, 23).value = f'=S{row}/100'
        ws_calc.cell(row, 23).font = black_font
        ws_calc.cell(row, 23).alignment = center_alignment
        ws_calc.cell(row, 23).border = thin_border
        ws_calc.cell(row, 23).number_format = '0.00'

        # X列: 达成度目标2 = T/100
        ws_calc.cell(row, 24).value = f'=T{row}/100'
        ws_calc.cell(row, 24).font = black_font
        ws_calc.cell(row, 24).alignment = center_alignment
        ws_calc.cell(row, 24).border = thin_border
        ws_calc.cell(row, 24).number_format = '0.00'

        # Y列: 达成度目标3 = U/100
        ws_calc.cell(row, 25).value = f'=U{row}/100'
        ws_calc.cell(row, 25).font = black_font
        ws_calc.cell(row, 25).alignment = center_alignment
        ws_calc.cell(row, 25).border = thin_border
        ws_calc.cell(row, 25).number_format = '0.00'

    # Z、AA、AB、AG列: 达成度平均值（每一行都填充相同的平均值，用于图表显示平均线）
    for row in range(data_start_row, data_end_row + 1):
        # Z列: 目标1达成度平均值
        ws_calc.cell(row, 26).value = f'=AVERAGE(W${data_start_row}:W${data_end_row})'
        ws_calc.cell(row, 26).font = black_font
        ws_calc.cell(row, 26).alignment = center_alignment
        ws_calc.cell(row, 26).border = thin_border
        ws_calc.cell(row, 26).number_format = '0.00'

        # AA列: 目标2达成度平均值
        ws_calc.cell(row, 27).value = f'=AVERAGE(X${data_start_row}:X${data_end_row})'
        ws_calc.cell(row, 27).font = black_font
        ws_calc.cell(row, 27).alignment = center_alignment
        ws_calc.cell(row, 27).border = thin_border
        ws_calc.cell(row, 27).number_format = '0.00'

        # AB列: 目标3达成度平均值
        ws_calc.cell(row, 28).value = f'=AVERAGE(Y${data_start_row}:Y${data_end_row})'
        ws_calc.cell(row, 28).font = black_font
        ws_calc.cell(row, 28).alignment = center_alignment
        ws_calc.cell(row, 28).border = thin_border
        ws_calc.cell(row, 28).number_format = '0.00'

    # AC-AE列: 达成度期望值（固定为0.6）
    for row in range(data_start_row, data_end_row + 1):
        ws_calc.cell(row, 29).value = 0.6  # AC列
        ws_calc.cell(row, 29).font = black_font
        ws_calc.cell(row, 29).alignment = center_alignment
        ws_calc.cell(row, 29).border = thin_border
        ws_calc.cell(row, 29).number_format = '0.00'

        ws_calc.cell(row, 30).value = 0.6  # AD列
        ws_calc.cell(row, 30).font = black_font
        ws_calc.cell(row, 30).alignment = center_alignment
        ws_calc.cell(row, 30).border = thin_border
        ws_calc.cell(row, 30).number_format = '0.00'

        ws_calc.cell(row, 31).value = 0.6  # AE列
        ws_calc.cell(row, 31).font = black_font
        ws_calc.cell(row, 31).alignment = center_alignment
        ws_calc.cell(row, 31).border = thin_border
        ws_calc.cell(row, 31).number_format = '0.00'

        # AF列: 总达成度 = V/100
        ws_calc.cell(row, 32).value = f'=V{row}/100'
        ws_calc.cell(row, 32).font = black_font
        ws_calc.cell(row, 32).alignment = center_alignment
        ws_calc.cell(row, 32).border = thin_border
        ws_calc.cell(row, 32).number_format = '0.00'

        # AG列: 总达成度平均值（每行都显示相同的平均值，用于图表平均线）
        ws_calc.cell(row, 33).value = f'=AVERAGE(AF${data_start_row}:AF${data_end_row})'
        ws_calc.cell(row, 33).font = black_font
        ws_calc.cell(row, 33).alignment = center_alignment
        ws_calc.cell(row, 33).border = thin_border
        ws_calc.cell(row, 33).number_format = '0.00'

    # 在平均值行合并A、B列单元格，标注"（平均值）"并居中
    ws_calc.merge_cells(f'A{avg_row}:B{avg_row}')
    ws_calc.cell(avg_row, 1).value = '（平均值）'
    ws_calc.cell(avg_row, 1).font = black_font
    ws_calc.cell(avg_row, 1).alignment = center_alignment
    ws_calc.cell(avg_row, 1).border = thin_border
    ws_calc.cell(avg_row, 2).border = thin_border  # B列也需要边框

    # 为所有数值列添加平均值
    # C-H列：目标得分和成绩
    for col in range(3, 9):  # C到H
        col_letter = get_column_letter(col)
        ws_calc.cell(avg_row, col).value = f'=AVERAGE({col_letter}{data_start_row}:{col_letter}{data_end_row})'
        ws_calc.cell(avg_row, col).font = black_font
        ws_calc.cell(avg_row, col).alignment = right_alignment
        ws_calc.cell(avg_row, col).border = thin_border
        ws_calc.cell(avg_row, col).number_format = '0.00'

    # K-V列：各种计算列
    for col in range(11, 23):  # K到V
        col_letter = get_column_letter(col)
        ws_calc.cell(avg_row, col).value = f'=AVERAGE({col_letter}{data_start_row}:{col_letter}{data_end_row})'
        ws_calc.cell(avg_row, col).font = black_font
        ws_calc.cell(avg_row, col).alignment = right_alignment
        ws_calc.cell(avg_row, col).border = thin_border
        ws_calc.cell(avg_row, col).number_format = '0.00'

    # W-Y列：达成度目标
    for col in range(23, 26):  # W到Y
        col_letter = get_column_letter(col)
        ws_calc.cell(avg_row, col).value = f'=AVERAGE({col_letter}{data_start_row}:{col_letter}{data_end_row})'
        ws_calc.cell(avg_row, col).font = black_font
        ws_calc.cell(avg_row, col).alignment = right_alignment
        ws_calc.cell(avg_row, col).border = thin_border
        ws_calc.cell(avg_row, col).number_format = '0.00'

    # AF列：总达成度
    ws_calc.cell(avg_row, 32).value = f'=AVERAGE(AF{data_start_row}:AF{data_end_row})'
    ws_calc.cell(avg_row, 32).font = black_font
    ws_calc.cell(avg_row, 32).alignment = right_alignment
    ws_calc.cell(avg_row, 32).border = thin_border
    ws_calc.cell(avg_row, 32).number_format = '0.00'

    # 清除多余的行数据（如果模板中有更多行的话）
    for row in range(avg_row + 1, avg_row + 100):
        for col in range(1, 34):  # A到AG列
            ws_calc.cell(row, col).value = None

    # 清除AH列（无用的计数列）- 清除值和边框
    no_border = Border()  # 无边框
    for row in range(1, avg_row + 100):
        ws_calc.cell(row, 34).value = None
        ws_calc.cell(row, 34).border = no_border  # 删除边框

    # 修改AF1的"算平均值"为"算术平均值"
    if ws_calc.cell(1, 32).value == '算平均值':
        ws_calc.cell(1, 32).value = '算术平均值'

    # 拆分并重设"平时成绩"和"期末成绩"的合并单元格
    # 取消原有合并
    try:
        ws_calc.unmerge_cells('K1:N1')
    except:
        pass
    try:
        ws_calc.unmerge_cells('O1:R1')
    except:
        pass

    # 重新合并：K1:L1（平时成绩标题）、M1:N1（平时比例）、O1:P1（期末成绩标题）、Q1:R1（期末比例）
    ws_calc.merge_cells('K1:L1')
    ws_calc.cell(1, 11).value = '平时成绩'
    ws_calc.cell(1, 11).font = bold_font  # 文本加粗
    ws_calc.cell(1, 11).alignment = center_alignment
    ws_calc.cell(1, 11).border = thin_border

    ws_calc.merge_cells('M1:N1')
    ws_calc.cell(1, 13).value = 30  # 平时成绩占比30%
    ws_calc.cell(1, 13).font = black_font  # 数值不加粗
    ws_calc.cell(1, 13).alignment = center_alignment
    ws_calc.cell(1, 13).border = thin_border

    ws_calc.merge_cells('O1:P1')
    ws_calc.cell(1, 15).value = '期末成绩'
    ws_calc.cell(1, 15).font = bold_font  # 文本加粗
    ws_calc.cell(1, 15).alignment = center_alignment
    ws_calc.cell(1, 15).border = thin_border

    ws_calc.merge_cells('Q1:R1')
    ws_calc.cell(1, 17).value = 70  # 期末成绩占比70%
    ws_calc.cell(1, 17).font = black_font  # 数值不加粗
    ws_calc.cell(1, 17).alignment = center_alignment
    ws_calc.cell(1, 17).border = thin_border

    # 修改S1:V1的"总成绩 改公式"为"总成绩"（删除提示文字）
    ws_calc.cell(1, 19).value = '总成绩'
    ws_calc.cell(1, 19).font = bold_font  # 文本加粗
    ws_calc.cell(1, 19).alignment = center_alignment

    # AG列"总达成度平均值"标题合并第1、2行单元格
    ws_calc.merge_cells('AG1:AG2')
    ws_calc.cell(1, 33).value = '总达成度平均值'
    ws_calc.cell(1, 33).font = bold_font  # 文本加粗
    ws_calc.cell(1, 33).alignment = center_alignment
    ws_calc.cell(1, 33).border = thin_border

    # 自动调整列宽以完整显示文本
    adjust_column_widths(ws_calc)
    adjust_column_widths(ws_stat)

    # 手动设置特定列的宽度
    # 数值列：5中文字符(10) + 两边各0.5字符 = 11
    numeric_width = 11
    # 数值列包括：C-H, K-Y, AF (除了序号I和姓名J)
    for col in range(3, 9):  # C-H
        ws_calc.column_dimensions[get_column_letter(col)].width = numeric_width
    for col in range(11, 26):  # K-Y
        ws_calc.column_dimensions[get_column_letter(col)].width = numeric_width
    for col in range(26, 29):  # Z-AB (平均值列)
        ws_calc.column_dimensions[get_column_letter(col)].width = numeric_width
    for col in range(29, 32):  # AC-AE (期望值列)
        ws_calc.column_dimensions[get_column_letter(col)].width = numeric_width
    ws_calc.column_dimensions['AF'].width = numeric_width + 2  # 总达成度，加宽1中文字符
    ws_calc.column_dimensions['AG'].width = 16.5  # 总达成度平均值，加宽1中文字符

    # 班级列：表演2211 = 8中文宽度 + 4数字 = 12 + 1边距 = 13
    ws_calc.column_dimensions['A'].width = 13
    # 学号列：12位数字 + 1边距 = 13
    ws_calc.column_dimensions['B'].width = 13
    # 姓名列：最多4中文 = 8 + 1边距 = 9
    ws_calc.column_dimensions['J'].width = 9
    # 序号列：加宽1中文字符
    ws_calc.column_dimensions['I'].width = 6

    # 更新达成度统计页
    update_statistics_sheet(ws_stat, num_students, data_start_row, data_end_row)

    # 设置达成度统计页的列宽和边框
    # A,B列：5中文字符 = 10 + 1边距 = 11
    ws_stat.column_dimensions['A'].width = 11
    ws_stat.column_dimensions['B'].width = 11
    # C-H列：3中文字符 = 6 + 1边距 = 7
    for col in ['C', 'D', 'E', 'F', 'G', 'H']:
        ws_stat.column_dimensions[col].width = 7

    # 为A1:H7所有单元格添加边框（包括合并单元格）
    for row in range(1, 8):
        for col in range(1, 9):  # A-H
            cell = ws_stat.cell(row, col)
            cell.border = thin_border
            cell.alignment = center_alignment
            # 保持已有字体，如果没有则设置默认字体
            if cell.font is None or cell.font.color is None:
                cell.font = black_font

    # 处理合并单元格的边框（C1:D1, E1:F1, G1:H1）
    # 合并单元格需要为每个角设置正确的边框
    merged_ranges = [('C1', 'D1'), ('E1', 'F1'), ('G1', 'H1')]
    for start, end in merged_ranges:
        start_col = ord(start[0]) - ord('A') + 1
        end_col = ord(end[0]) - ord('A') + 1
        row_num = int(start[1])
        # 左边单元格：左、上、下边框
        ws_stat.cell(row_num, start_col).border = Border(
            left=Side(style='thin'),
            right=Side(style=None),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        # 右边单元格：右、上、下边框
        ws_stat.cell(row_num, end_col).border = Border(
            left=Side(style=None),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # 更新图表
    update_charts(wb, ws_calc, ws_stat, num_students, data_start_row, data_end_row)

    # 对齐图表
    align_charts(ws_calc, ws_stat)

    # 保存输出文件
    wb.save(output_file)
    print(f"输出文件已保存: {output_file}")


def adjust_column_widths(ws):
    """自动调整列宽以完整显示文本"""
    from openpyxl.utils import get_column_letter

    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)

        for row_idx in range(1, min(ws.max_row + 1, 100)):  # 检查前100行
            cell = ws.cell(row_idx, col_idx)
            if cell.value:
                cell_value = str(cell.value)

                # 跳过公式，只考虑实际显示的值
                if cell_value.startswith('='):
                    # 对于公式，估算显示宽度（数值通常6-8个字符）
                    cell_length = 8
                else:
                    # 计算单元格内容长度
                    cell_length = 0
                    for char in cell_value:
                        if '\u4e00' <= char <= '\u9fff':
                            cell_length += 2
                        else:
                            cell_length += 1
                max_length = max(max_length, cell_length)

        # 设置列宽（添加小边距）
        if max_length > 0:
            adjusted_width = min(max_length + 0.5, 50)  # 最小边距，最大宽度50
            ws.column_dimensions[column_letter].width = adjusted_width


def update_statistics_sheet(ws_stat, num_students, data_start_row, data_end_row):
    """更新达成度统计页"""
    black_font = Font(color="000000")
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 更新占比公式中的总人数
    for row in range(3, 8):
        # D列: 目标1占比
        ws_stat.cell(row, 4).value = f'=C{row}/{num_students}'
        ws_stat.cell(row, 4).font = black_font
        ws_stat.cell(row, 4).alignment = center_alignment
        ws_stat.cell(row, 4).border = thin_border
        ws_stat.cell(row, 4).number_format = '0.00%'

        # F列: 目标2占比
        ws_stat.cell(row, 6).value = f'=E{row}/{num_students}'
        ws_stat.cell(row, 6).font = black_font
        ws_stat.cell(row, 6).alignment = center_alignment
        ws_stat.cell(row, 6).border = thin_border
        ws_stat.cell(row, 6).number_format = '0.00%'

        # H列: 目标3占比
        ws_stat.cell(row, 8).value = f'=G{row}/{num_students}'
        ws_stat.cell(row, 8).font = black_font
        ws_stat.cell(row, 8).alignment = center_alignment
        ws_stat.cell(row, 8).border = thin_border
        ws_stat.cell(row, 8).number_format = '0.00%'

    # 人数统计使用COUNTIF公式
    # 基于达成度列(W, X, Y)进行统计
    # 完全达成: >0.8
    # 较好达成: 0.6-0.8
    # 基本达成: 0.5-0.6
    # 较少达成: 0.4-0.5
    # 没有达成: <0.4

    # 目标1人数统计 (基于W列 - 达成度目标1)
    ws_stat.cell(3, 3).value = f'=COUNTIF(\'课程目标达成度计算\'!W{data_start_row}:W{data_end_row},">0.8")'
    ws_stat.cell(4, 3).value = f'=COUNTIFS(\'课程目标达成度计算\'!W{data_start_row}:W{data_end_row},">=0.6",\'课程目标达成度计算\'!W{data_start_row}:W{data_end_row},"<=0.8")'
    ws_stat.cell(5, 3).value = f'=COUNTIFS(\'课程目标达成度计算\'!W{data_start_row}:W{data_end_row},">=0.5",\'课程目标达成度计算\'!W{data_start_row}:W{data_end_row},"<0.6")'
    ws_stat.cell(6, 3).value = f'=COUNTIFS(\'课程目标达成度计算\'!W{data_start_row}:W{data_end_row},">=0.4",\'课程目标达成度计算\'!W{data_start_row}:W{data_end_row},"<0.5")'
    ws_stat.cell(7, 3).value = f'=COUNTIF(\'课程目标达成度计算\'!W{data_start_row}:W{data_end_row},"<0.4")'

    # 目标2人数统计 (基于X列 - 达成度目标2)
    ws_stat.cell(3, 5).value = f'=COUNTIF(\'课程目标达成度计算\'!X{data_start_row}:X{data_end_row},">0.8")'
    ws_stat.cell(4, 5).value = f'=COUNTIFS(\'课程目标达成度计算\'!X{data_start_row}:X{data_end_row},">=0.6",\'课程目标达成度计算\'!X{data_start_row}:X{data_end_row},"<=0.8")'
    ws_stat.cell(5, 5).value = f'=COUNTIFS(\'课程目标达成度计算\'!X{data_start_row}:X{data_end_row},">=0.5",\'课程目标达成度计算\'!X{data_start_row}:X{data_end_row},"<0.6")'
    ws_stat.cell(6, 5).value = f'=COUNTIFS(\'课程目标达成度计算\'!X{data_start_row}:X{data_end_row},">=0.4",\'课程目标达成度计算\'!X{data_start_row}:X{data_end_row},"<0.5")'
    ws_stat.cell(7, 5).value = f'=COUNTIF(\'课程目标达成度计算\'!X{data_start_row}:X{data_end_row},"<0.4")'

    # 目标3人数统计 (基于Y列 - 达成度目标3)
    ws_stat.cell(3, 7).value = f'=COUNTIF(\'课程目标达成度计算\'!Y{data_start_row}:Y{data_end_row},">0.8")'
    ws_stat.cell(4, 7).value = f'=COUNTIFS(\'课程目标达成度计算\'!Y{data_start_row}:Y{data_end_row},">=0.6",\'课程目标达成度计算\'!Y{data_start_row}:Y{data_end_row},"<=0.8")'
    ws_stat.cell(5, 7).value = f'=COUNTIFS(\'课程目标达成度计算\'!Y{data_start_row}:Y{data_end_row},">=0.5",\'课程目标达成度计算\'!Y{data_start_row}:Y{data_end_row},"<0.6")'
    ws_stat.cell(6, 7).value = f'=COUNTIFS(\'课程目标达成度计算\'!Y{data_start_row}:Y{data_end_row},">=0.4",\'课程目标达成度计算\'!Y{data_start_row}:Y{data_end_row},"<0.5")'
    ws_stat.cell(7, 7).value = f'=COUNTIF(\'课程目标达成度计算\'!Y{data_start_row}:Y{data_end_row},"<0.4")'

    # 设置样式（包括人数列C、E、G）
    for row in range(3, 8):
        for col in [3, 5, 7]:  # 人数列
            ws_stat.cell(row, col).font = black_font
            ws_stat.cell(row, col).alignment = center_alignment
            ws_stat.cell(row, col).border = thin_border


def update_charts(wb, ws_calc, ws_stat, num_students, data_start_row, data_end_row):
    """更新图表数据范围和横坐标轴边界"""
    from openpyxl.chart.data_source import NumRef, NumData, AxDataSource, NumDataSource

    # 更新课程目标达成度计算页的散点图
    for chart in ws_calc._charts:
        if isinstance(chart, ScatterChart):
            try:
                # 更新X轴范围
                if chart.x_axis.scaling is None:
                    chart.x_axis.scaling = Scaling()
                chart.x_axis.scaling.min = 0
                chart.x_axis.scaling.max = num_students + 5

                # 更新数据系列范围
                for series in chart.series:
                    # 更新xVal范围
                    if hasattr(series, 'xVal') and series.xVal:
                        if hasattr(series.xVal, 'numRef') and series.xVal.numRef:
                            old_ref = series.xVal.numRef.f
                            # 替换行范围
                            new_ref = update_range_reference(old_ref, data_start_row, data_end_row)
                            series.xVal.numRef.f = new_ref

                    # 更新yVal范围
                    if hasattr(series, 'yVal') and series.yVal:
                        if hasattr(series.yVal, 'numRef') and series.yVal.numRef:
                            old_ref = series.yVal.numRef.f
                            new_ref = update_range_reference(old_ref, data_start_row, data_end_row)
                            series.yVal.numRef.f = new_ref

            except Exception as e:
                print(f"警告: 更新散点图时出错: {e}")

    print(f"图表已更新: X轴最大值设为 {num_students + 5}")


def update_range_reference(ref, start_row, end_row):
    """更新单元格范围引用中的行号"""
    import re
    # 匹配类似 $I$3:$I$89 的模式
    pattern = r'(\$[A-Z]+\$)\d+:(\$[A-Z]+\$)\d+'
    replacement = f'\\g<1>{start_row}:\\g<2>{end_row}'
    return re.sub(pattern, replacement, ref)


def align_charts(ws_calc, ws_stat):
    """对齐图表，删除无用柱状图，使其整齐排列"""
    from openpyxl.chart.text import RichText
    from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font as DrawingFont

    # 课程目标达成度计算页的图表处理
    # 删除柱状图(图表0)，只保留4个散点图
    # 图表布局: 2行2列
    # 第一行: 目标1散点图, 目标2散点图
    # 第二行: 目标3散点图, 总达成度散点图

    charts = ws_calc._charts

    # 删除柱状图（索引0的BarChart）
    if len(charts) >= 5:
        # 找到并删除柱状图
        from openpyxl.chart import BarChart
        bar_chart = None
        for chart in charts:
            if isinstance(chart, BarChart):
                bar_chart = chart
                break
        if bar_chart:
            ws_calc._charts.remove(bar_chart)
            print("已删除无用的柱状图")

    # 重新获取图表列表（删除后）
    charts = ws_calc._charts
    if len(charts) >= 4:
        # 定义图表尺寸（增大以提高分辨率，单位为厘米）
        chart_width = 18  # 图表宽度（厘米）
        chart_height = 12  # 图表高度（厘米）

        # 图表间距（AJ列=36，AI列=35空出来作为间距）
        col_gap = 13  # 左右间距（列数，约等于图表宽度）
        row_gap = 24  # 上下间距（行数）

        # 第一行起始位置（从AJ=36列开始，与第2行对齐）
        start_col = 36  # AJ列
        row1_start = 2  # 与第2行对齐

        for i, chart in enumerate(charts):
            # 计算位置
            col_offset = (i % 2) * col_gap
            row_offset = (i // 2) * row_gap
            chart.anchor = f'{get_column_letter(start_col + col_offset)}{row1_start + row_offset}'

            # 设置图表尺寸（必须在设置anchor之后）
            chart.width = chart_width
            chart.height = chart_height

    # 达成度统计页的图表对齐
    stat_charts = ws_stat._charts
    if len(stat_charts) >= 3:
        # 三个柱状图水平排列，从A列开始（靠最左）
        stat_chart_width = 12  # 加宽图表使标题能在一行内显示
        stat_chart_height = 10
        stat_start_col = 1  # A列开始（靠最左）
        stat_start_row = 9  # 数据在1-7行，图表从第9行开始

        for i, chart in enumerate(stat_charts):
            chart.width = stat_chart_width
            chart.height = stat_chart_height
            chart.anchor = f'{get_column_letter(stat_start_col + i * 9)}{stat_start_row}'

            # 设置X轴标签不旋转（水平显示）
            if hasattr(chart, 'x_axis') and chart.x_axis is not None:
                from openpyxl.chart.text import RichText
                from openpyxl.drawing.text import RichTextProperties, Paragraph, ParagraphProperties, CharacterProperties
                # 创建新的文本属性，旋转角度为0
                chart.x_axis.txPr = RichText(
                    bodyPr=RichTextProperties(rot=0),
                    p=[Paragraph(
                        pPr=ParagraphProperties(
                            defRPr=CharacterProperties(sz=900)
                        )
                    )]
                )

    print("图表已对齐")


def main():
    """主函数"""
    grades_file = '2022-2023第一学期总评成绩(按行政班).xlsx'
    template_file = '达成度数据模板.xlsx'
    output_file = '达成度数据输出.xlsx'

    print("=" * 50)
    print("达成度数据处理脚本")
    print("=" * 50)

    # 1. 提取学生数据
    print("\n[1/3] 从成绩文件提取学生数据...")
    students = extract_students_from_grades(grades_file)
    print(f"成功提取 {len(students)} 名学生数据")

    # 2. 排序
    print("\n[2/3] 按行政班分组，按学号升序排序...")
    students = sort_students(students)

    # 显示排序后的班级统计
    from collections import Counter
    classes = Counter([s['class'] for s in students])
    for cls, count in sorted(classes.items()):
        print(f"  {cls}: {count}人")

    # 3. 处理模板并输出
    print("\n[3/3] 处理模板文件...")
    process_template(template_file, output_file, students)

    print("\n" + "=" * 50)
    print("处理完成！")
    print("=" * 50)


if __name__ == '__main__':
    main()
