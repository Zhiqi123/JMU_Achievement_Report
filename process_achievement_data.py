#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
达成度数据处理脚本
从成绩数据生成达成度报告Excel文件（完全独立，不依赖模板）
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.chart import BarChart
from openpyxl.utils import get_column_letter


# ==================== 配置参数 ====================
# 达成度目标占比（可根据需要修改）
RATIO_1 = 50  # 目标一占比 (%)
RATIO_2 = 30  # 目标二占比 (%)
RATIO_3 = 20  # 目标三占比 (%)

# 成绩占比
REGULAR_SCORE_RATIO = 30  # 平时成绩占比 (%)
FINAL_SCORE_RATIO = 70    # 期末成绩占比 (%)

# 达成度期望值
ACHIEVEMENT_EXPECTATION = 0.6
# ================================================


def extract_students_from_grades(grades_file):
    """从成绩文件中提取所有学生数据（动态识别列结构）"""
    import re

    xl = pd.ExcelFile(grades_file)
    all_students = []

    for sheet in xl.sheet_names:
        if sheet == 'Sheet1':
            continue

        df = pd.read_excel(xl, sheet_name=sheet, header=None)

        # ===== 1. 动态查找行政班信息 =====
        class_name = None
        for i in range(min(10, len(df))):  # 在前10行中搜索
            for j in range(min(5, len(df.columns))):  # 在前5列中搜索
                cell_value = str(df.iloc[i, j]) if pd.notna(df.iloc[i, j]) else ''
                if '行政班' in cell_value:
                    # 提取行政班名称，处理格式如"行政班：音乐2212(音乐2212)  授课教师：范小龙"
                    match = re.search(r'行政班[：:]\s*([^\s(（]+)', cell_value)
                    if match:
                        class_name = match.group(1).strip()
                    break
            if class_name:
                break

        if not class_name:
            continue  # 未找到行政班信息，跳过此工作表

        # ===== 2. 动态查找列头行 =====
        header_row = None
        col_mapping = {}  # 存储列名到列索引的映射

        # 定义要搜索的关键字及其可能的变体
        key_patterns = {
            'student_id': ['学号'],
            'name': ['姓名'],
            'final_score': ['期末成绩', '期末', '期末考试'],
            'regular_score': ['平时成绩', '平时', '平时分'],
            'total_score': ['总成绩', '总评成绩', '成绩', '总评']  # 优先级从高到低
        }

        for i in range(min(15, len(df))):  # 在前15行中搜索列头
            row_values = [str(df.iloc[i, j]).strip() if pd.notna(df.iloc[i, j]) else '' for j in range(len(df.columns))]

            # 检查是否包含"学号"和"姓名"（这是列头行的标志）
            if any('学号' in v for v in row_values) and any('姓名' in v for v in row_values):
                header_row = i

                # 建立列名到索引的映射
                for j, cell_value in enumerate(row_values):
                    cell_value = cell_value.strip()

                    # 学号
                    if '学号' in cell_value and 'student_id' not in col_mapping:
                        col_mapping['student_id'] = j

                    # 姓名
                    if '姓名' in cell_value and 'name' not in col_mapping:
                        col_mapping['name'] = j

                    # 期末成绩
                    if any(p in cell_value for p in key_patterns['final_score']) and 'final_score' not in col_mapping:
                        col_mapping['final_score'] = j

                    # 平时成绩
                    if any(p in cell_value for p in key_patterns['regular_score']) and 'regular_score' not in col_mapping:
                        col_mapping['regular_score'] = j

                    # 总成绩（优先匹配"总成绩"、"总评成绩"，其次匹配单独的"成绩"）
                    if 'total_score' not in col_mapping:
                        if '总成绩' in cell_value or '总评成绩' in cell_value:
                            col_mapping['total_score'] = j
                        elif cell_value == '成绩' or cell_value == '总评':
                            # 单独的"成绩"作为备选
                            col_mapping['total_score'] = j

                break

        if header_row is None:
            continue  # 未找到列头行，跳过此工作表

        # 检查是否找到了所有必需的列
        required_cols = ['student_id', 'name', 'final_score', 'regular_score', 'total_score']
        missing_cols = [col for col in required_cols if col not in col_mapping]
        if missing_cols:
            print(f"  警告: 工作表 {sheet} 缺少列: {missing_cols}，跳过")
            continue

        # ===== 3. 提取学生数据 =====
        data_start_row = header_row + 1

        for i in range(data_start_row, len(df)):
            row = df.iloc[i]

            # 获取学号
            student_id = str(row[col_mapping['student_id']]) if pd.notna(row[col_mapping['student_id']]) else ''

            # 检查是否是有效学生数据行（学号为纯数字且长度大于8）
            if student_id.isdigit() and len(student_id) > 8:
                name = row[col_mapping['name']]

                # 获取各项成绩，检测缺考/缓考等特殊状态
                final_raw = row[col_mapping['final_score']]
                regular_raw = row[col_mapping['regular_score']]
                total_raw = row[col_mapping['total_score']]

                # 检测特殊状态（缺考、缓考等）
                special_status = None
                special_keywords = ['缺考', '缓考', '作弊', '取消', '免修', '旷考']

                for raw_val in [final_raw, regular_raw, total_raw]:
                    if pd.notna(raw_val):
                        raw_str = str(raw_val).strip()
                        for keyword in special_keywords:
                            if keyword in raw_str:
                                special_status = raw_str
                                break
                    if special_status:
                        break

                # 检查是否所有成绩都为空（使用标量安全的检查方式）
                def is_empty(val):
                    """检查单个值是否为空"""
                    if val is None:
                        return True
                    try:
                        if pd.isna(val):
                            return True
                    except (ValueError, TypeError):
                        pass
                    return str(val).strip() == ''

                all_empty = is_empty(final_raw) and is_empty(regular_raw) and is_empty(total_raw)

                if all_empty:
                    special_status = '成绩为空'

                if special_status:
                    # 特殊状态学生：保留基本信息，标记状态
                    all_students.append({
                        'class': class_name,
                        'student_id': student_id,
                        'name': name,
                        'final_score': None,
                        'regular_score': None,
                        'total_score': None,
                        'status': special_status  # 特殊状态标记
                    })
                else:
                    # 正常学生：尝试转换成绩
                    try:
                        final_score = float(final_raw)
                        regular_score = float(regular_raw)
                        total_score = float(total_raw)

                        all_students.append({
                            'class': class_name,
                            'student_id': student_id,
                            'name': name,
                            'final_score': final_score,
                            'regular_score': regular_score,
                            'total_score': total_score,
                            'status': None  # 正常状态
                        })
                    except (ValueError, TypeError):
                        # 成绩格式异常，标记为特殊状态
                        all_students.append({
                            'class': class_name,
                            'student_id': student_id,
                            'name': name,
                            'final_score': None,
                            'regular_score': None,
                            'total_score': None,
                            'status': '成绩异常'
                        })

    return all_students


def sort_students(students):
    """按行政班分组，按学号升序排序"""
    # 先按班级排序，再按学号排序
    return sorted(students, key=lambda x: (x['class'], x['student_id']))


def create_workbook(output_file, students):
    """从零创建工作簿，填入学生数据并生成输出文件"""

    # 创建新工作簿
    wb = openpyxl.Workbook()

    # 删除默认的Sheet，创建两个工作表
    default_sheet = wb.active
    ws_calc = wb.create_sheet('课程目标达成度计算', 0)
    ws_stat = wb.create_sheet('达成度统计', 1)
    wb.remove(default_sheet)

    # 使用配置的占比值
    ratio_1 = RATIO_1
    ratio_2 = RATIO_2
    ratio_3 = RATIO_3

    print(f"达成度占比: 目标一={ratio_1}%, 目标二={ratio_2}%, 目标三={ratio_3}%")

    # 定义样式
    black_font = Font(color="000000")
    bold_font = Font(color="000000", bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    right_alignment = Alignment(horizontal='right', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 计算需要的行数
    num_students = len(students)
    data_start_row = 3
    data_end_row = data_start_row + num_students - 1
    avg_row = data_end_row + 1  # 平均值行

    print(f"学生数量: {num_students}")
    print(f"数据行: {data_start_row} - {data_end_row}")
    print(f"平均值行: {avg_row}")

    # ==================== 创建第一行标题 ====================
    setup_calc_sheet_headers(ws_calc, ratio_1, ratio_2, ratio_3, bold_font, black_font, center_alignment, thin_border)

    # ==================== 填入学生数据 ====================
    for idx, student in enumerate(students):
        row = data_start_row + idx
        is_special = student.get('status') is not None  # 是否为特殊状态学生

        # A列: 班级
        ws_calc.cell(row, 1).value = student['class']
        ws_calc.cell(row, 1).font = black_font
        ws_calc.cell(row, 1).alignment = center_alignment
        ws_calc.cell(row, 1).border = thin_border

        # B列: 学号
        ws_calc.cell(row, 2).value = student['student_id']
        ws_calc.cell(row, 2).font = black_font
        ws_calc.cell(row, 2).alignment = center_alignment
        ws_calc.cell(row, 2).border = thin_border

        # I列: 序号（从1开始）- 无论是否特殊状态都写入
        ws_calc.cell(row, 9).value = idx + 1
        ws_calc.cell(row, 9).font = black_font
        ws_calc.cell(row, 9).alignment = center_alignment
        ws_calc.cell(row, 9).border = thin_border

        # J列: 姓名
        ws_calc.cell(row, 10).value = student['name']
        ws_calc.cell(row, 10).font = black_font
        ws_calc.cell(row, 10).alignment = center_alignment
        ws_calc.cell(row, 10).border = thin_border

        if is_special:
            # 特殊状态学生：只写入基本信息，H列显示状态，其他列留空（只设置边框）
            # C-E列: 目标分数 - 留空
            for col in range(3, 6):
                ws_calc.cell(row, col).border = thin_border

            # F列: 平时成绩 - 留空
            ws_calc.cell(row, 6).border = thin_border

            # G列: 期末成绩 - 留空
            ws_calc.cell(row, 7).border = thin_border

            # H列: 总成绩 - 显示特殊状态
            ws_calc.cell(row, 8).value = student['status']
            ws_calc.cell(row, 8).font = black_font
            ws_calc.cell(row, 8).alignment = center_alignment
            ws_calc.cell(row, 8).border = thin_border

            # K-Y列: 达成度相关 - 留空
            for col in range(11, 26):
                ws_calc.cell(row, col).border = thin_border

        else:
            # 正常学生：写入所有数据和公式
            # C列: 目标一 = ROUND(总成绩 * $C$1 / 100, 0)
            ws_calc.cell(row, 3).value = f'=ROUND(H{row}*$C$1/100,0)'
            ws_calc.cell(row, 3).font = black_font
            ws_calc.cell(row, 3).alignment = center_alignment
            ws_calc.cell(row, 3).border = thin_border

            # D列: 目标二 = ROUND(总成绩 * $D$1 / 100, 0)
            ws_calc.cell(row, 4).value = f'=ROUND(H{row}*$D$1/100,0)'
            ws_calc.cell(row, 4).font = black_font
            ws_calc.cell(row, 4).alignment = center_alignment
            ws_calc.cell(row, 4).border = thin_border

            # E列: 目标三 = ROUND(总成绩 * $E$1 / 100, 0)
            ws_calc.cell(row, 5).value = f'=ROUND(H{row}*$E$1/100,0)'
            ws_calc.cell(row, 5).font = black_font
            ws_calc.cell(row, 5).alignment = center_alignment
            ws_calc.cell(row, 5).border = thin_border

            # F列: 平时成绩
            ws_calc.cell(row, 6).value = student['regular_score']
            ws_calc.cell(row, 6).font = black_font
            ws_calc.cell(row, 6).alignment = center_alignment
            ws_calc.cell(row, 6).border = thin_border
            ws_calc.cell(row, 6).number_format = '0.00'

            # G列: 期末成绩
            ws_calc.cell(row, 7).value = student['final_score']
            ws_calc.cell(row, 7).font = black_font
            ws_calc.cell(row, 7).alignment = center_alignment
            ws_calc.cell(row, 7).border = thin_border
            ws_calc.cell(row, 7).number_format = '0.00'

            # H列: 总成绩
            ws_calc.cell(row, 8).value = student['total_score']
            ws_calc.cell(row, 8).font = black_font
            ws_calc.cell(row, 8).alignment = center_alignment
            ws_calc.cell(row, 8).border = thin_border
            ws_calc.cell(row, 8).number_format = '0.00'

            # K列: 平时成绩目标1达成率
            ws_calc.cell(row, 11).value = f'=(ROUND(F{row}*$C$1/100,0)/$C$1)*100'
            ws_calc.cell(row, 11).font = black_font
            ws_calc.cell(row, 11).alignment = center_alignment
            ws_calc.cell(row, 11).border = thin_border
            ws_calc.cell(row, 11).number_format = '0.00'

            # L列: 平时成绩目标2达成率
            ws_calc.cell(row, 12).value = f'=(ROUND(F{row}*$D$1/100,0)/$D$1)*100'
            ws_calc.cell(row, 12).font = black_font
            ws_calc.cell(row, 12).alignment = center_alignment
            ws_calc.cell(row, 12).border = thin_border
            ws_calc.cell(row, 12).number_format = '0.00'

            # M列: 平时成绩目标3达成率
            ws_calc.cell(row, 13).value = f'=(ROUND(F{row}*$E$1/100,0)/$E$1)*100'
            ws_calc.cell(row, 13).font = black_font
            ws_calc.cell(row, 13).alignment = center_alignment
            ws_calc.cell(row, 13).border = thin_border
            ws_calc.cell(row, 13).number_format = '0.00'

            # N列: 平时成绩 = F列原值
            ws_calc.cell(row, 14).value = f'=F{row}'
            ws_calc.cell(row, 14).font = black_font
            ws_calc.cell(row, 14).alignment = center_alignment
            ws_calc.cell(row, 14).border = thin_border
            ws_calc.cell(row, 14).number_format = '0.00'

            # O列: 期末成绩目标1达成率
            ws_calc.cell(row, 15).value = f'=(ROUND(G{row}*$C$1/100,0)/$C$1)*100'
            ws_calc.cell(row, 15).font = black_font
            ws_calc.cell(row, 15).alignment = center_alignment
            ws_calc.cell(row, 15).border = thin_border
            ws_calc.cell(row, 15).number_format = '0.00'

            # P列: 期末成绩目标2达成率
            ws_calc.cell(row, 16).value = f'=(ROUND(G{row}*$D$1/100,0)/$D$1)*100'
            ws_calc.cell(row, 16).font = black_font
            ws_calc.cell(row, 16).alignment = center_alignment
            ws_calc.cell(row, 16).border = thin_border
            ws_calc.cell(row, 16).number_format = '0.00'

            # Q列: 期末成绩目标3达成率
            ws_calc.cell(row, 17).value = f'=(ROUND(G{row}*$E$1/100,0)/$E$1)*100'
            ws_calc.cell(row, 17).font = black_font
            ws_calc.cell(row, 17).alignment = center_alignment
            ws_calc.cell(row, 17).border = thin_border
            ws_calc.cell(row, 17).number_format = '0.00'

            # R列: 期末成绩 = G列原值
            ws_calc.cell(row, 18).value = f'=G{row}'
            ws_calc.cell(row, 18).font = black_font
            ws_calc.cell(row, 18).alignment = center_alignment
            ws_calc.cell(row, 18).border = thin_border
            ws_calc.cell(row, 18).number_format = '0.00'

            # S列: 总成绩目标1 = K*平时比例+O*期末比例
            ws_calc.cell(row, 19).value = f'=K{row}*$M$1/100+O{row}*$Q$1/100'
            ws_calc.cell(row, 19).font = black_font
            ws_calc.cell(row, 19).alignment = center_alignment
            ws_calc.cell(row, 19).border = thin_border
            ws_calc.cell(row, 19).number_format = '0.00'

            # T列: 总成绩目标2 = L*平时比例+P*期末比例
            ws_calc.cell(row, 20).value = f'=L{row}*$M$1/100+P{row}*$Q$1/100'
            ws_calc.cell(row, 20).font = black_font
            ws_calc.cell(row, 20).alignment = center_alignment
            ws_calc.cell(row, 20).border = thin_border
            ws_calc.cell(row, 20).number_format = '0.00'

            # U列: 总成绩目标3 = M*平时比例+Q*期末比例
            ws_calc.cell(row, 21).value = f'=M{row}*$M$1/100+Q{row}*$Q$1/100'
            ws_calc.cell(row, 21).font = black_font
            ws_calc.cell(row, 21).alignment = center_alignment
            ws_calc.cell(row, 21).border = thin_border
            ws_calc.cell(row, 21).number_format = '0.00'

            # V列: 总成绩 = H列
            ws_calc.cell(row, 22).value = f'=H{row}'
            ws_calc.cell(row, 22).font = black_font
            ws_calc.cell(row, 22).alignment = center_alignment
            ws_calc.cell(row, 22).border = thin_border
            ws_calc.cell(row, 22).number_format = '0.00'

            # W列: 达成度目标1 = S/100
            ws_calc.cell(row, 23).value = f'=S{row}/100'
            ws_calc.cell(row, 23).font = black_font
            ws_calc.cell(row, 23).alignment = center_alignment
            ws_calc.cell(row, 23).border = thin_border
            ws_calc.cell(row, 23).number_format = '0.00'

            # X列: 达成度目标2 = T/100
            ws_calc.cell(row, 24).value = f'=T{row}/100'
            ws_calc.cell(row, 24).font = black_font
            ws_calc.cell(row, 24).alignment = center_alignment
            ws_calc.cell(row, 24).border = thin_border
            ws_calc.cell(row, 24).number_format = '0.00'

            # Y列: 达成度目标3 = U/100
            ws_calc.cell(row, 25).value = f'=U{row}/100'
            ws_calc.cell(row, 25).font = black_font
            ws_calc.cell(row, 25).alignment = center_alignment
            ws_calc.cell(row, 25).border = thin_border
            ws_calc.cell(row, 25).number_format = '0.00'

    # Z、AA、AB、AC、AD、AE、AF、AG列: 根据学生状态设置值或仅设置边框
    for idx, student in enumerate(students):
        row = data_start_row + idx
        is_special = student.get('status') is not None

        # AC-AE列: 达成度期望值（所有学生都填入，保证图表红色虚线完整）
        ws_calc.cell(row, 29).value = ACHIEVEMENT_EXPECTATION  # AC列
        ws_calc.cell(row, 29).font = black_font
        ws_calc.cell(row, 29).alignment = center_alignment
        ws_calc.cell(row, 29).border = thin_border
        ws_calc.cell(row, 29).number_format = '0.00'

        ws_calc.cell(row, 30).value = ACHIEVEMENT_EXPECTATION  # AD列
        ws_calc.cell(row, 30).font = black_font
        ws_calc.cell(row, 30).alignment = center_alignment
        ws_calc.cell(row, 30).border = thin_border
        ws_calc.cell(row, 30).number_format = '0.00'

        ws_calc.cell(row, 31).value = ACHIEVEMENT_EXPECTATION  # AE列
        ws_calc.cell(row, 31).font = black_font
        ws_calc.cell(row, 31).alignment = center_alignment
        ws_calc.cell(row, 31).border = thin_border
        ws_calc.cell(row, 31).number_format = '0.00'

        if is_special:
            # 特殊状态学生：Z-AB、AF-AG列留空（只设置边框）
            for col in [26, 27, 28, 32, 33]:  # Z, AA, AB, AF, AG
                ws_calc.cell(row, col).border = thin_border
        else:
            # 正常学生：写入公式和值
            # Z列: 目标1达成度平均值
            ws_calc.cell(row, 26).value = f'=AVERAGE(W${data_start_row}:W${data_end_row})'
            ws_calc.cell(row, 26).font = black_font
            ws_calc.cell(row, 26).alignment = center_alignment
            ws_calc.cell(row, 26).border = thin_border
            ws_calc.cell(row, 26).number_format = '0.00'

            # AA列: 目标2达成度平均值
            ws_calc.cell(row, 27).value = f'=AVERAGE(X${data_start_row}:X${data_end_row})'
            ws_calc.cell(row, 27).font = black_font
            ws_calc.cell(row, 27).alignment = center_alignment
            ws_calc.cell(row, 27).border = thin_border
            ws_calc.cell(row, 27).number_format = '0.00'

            # AB列: 目标3达成度平均值
            ws_calc.cell(row, 28).value = f'=AVERAGE(Y${data_start_row}:Y${data_end_row})'
            ws_calc.cell(row, 28).font = black_font
            ws_calc.cell(row, 28).alignment = center_alignment
            ws_calc.cell(row, 28).border = thin_border
            ws_calc.cell(row, 28).number_format = '0.00'

            # AF列: 总达成度 = V/100
            ws_calc.cell(row, 32).value = f'=V{row}/100'
            ws_calc.cell(row, 32).font = black_font
            ws_calc.cell(row, 32).alignment = center_alignment
            ws_calc.cell(row, 32).border = thin_border
            ws_calc.cell(row, 32).number_format = '0.00'

            # AG列: 总达成度平均值
            ws_calc.cell(row, 33).value = f'=AVERAGE(AF${data_start_row}:AF${data_end_row})'
            ws_calc.cell(row, 33).font = black_font
            ws_calc.cell(row, 33).alignment = center_alignment
            ws_calc.cell(row, 33).border = thin_border
            ws_calc.cell(row, 33).number_format = '0.00'

    # 在平均值行合并A、B列单元格
    ws_calc.merge_cells(f'A{avg_row}:B{avg_row}')
    ws_calc.cell(avg_row, 1).value = '（平均值）'
    ws_calc.cell(avg_row, 1).font = black_font
    ws_calc.cell(avg_row, 1).alignment = center_alignment
    ws_calc.cell(avg_row, 1).border = thin_border
    ws_calc.cell(avg_row, 2).border = thin_border

    # 为所有数值列添加平均值
    # C-H列
    for col in range(3, 9):
        col_letter = get_column_letter(col)
        ws_calc.cell(avg_row, col).value = f'=AVERAGE({col_letter}{data_start_row}:{col_letter}{data_end_row})'
        ws_calc.cell(avg_row, col).font = black_font
        ws_calc.cell(avg_row, col).alignment = right_alignment
        ws_calc.cell(avg_row, col).border = thin_border
        ws_calc.cell(avg_row, col).number_format = '0.00'

    # K-V列
    for col in range(11, 23):
        col_letter = get_column_letter(col)
        ws_calc.cell(avg_row, col).value = f'=AVERAGE({col_letter}{data_start_row}:{col_letter}{data_end_row})'
        ws_calc.cell(avg_row, col).font = black_font
        ws_calc.cell(avg_row, col).alignment = right_alignment
        ws_calc.cell(avg_row, col).border = thin_border
        ws_calc.cell(avg_row, col).number_format = '0.00'

    # W-Y列
    for col in range(23, 26):
        col_letter = get_column_letter(col)
        ws_calc.cell(avg_row, col).value = f'=AVERAGE({col_letter}{data_start_row}:{col_letter}{data_end_row})'
        ws_calc.cell(avg_row, col).font = black_font
        ws_calc.cell(avg_row, col).alignment = right_alignment
        ws_calc.cell(avg_row, col).border = thin_border
        ws_calc.cell(avg_row, col).number_format = '0.00'

    # AF列
    ws_calc.cell(avg_row, 32).value = f'=AVERAGE(AF{data_start_row}:AF{data_end_row})'
    ws_calc.cell(avg_row, 32).font = black_font
    ws_calc.cell(avg_row, 32).alignment = right_alignment
    ws_calc.cell(avg_row, 32).border = thin_border
    ws_calc.cell(avg_row, 32).number_format = '0.00'

    # 设置列宽
    setup_column_widths(ws_calc)

    # 创建达成度统计页
    setup_statistics_sheet(ws_stat, data_start_row, data_end_row, black_font, bold_font, center_alignment, thin_border)

    # 为图表设置数据范围
    chart_start_row = data_start_row  # 图表数据起始行
    chart_end_row = data_end_row  # 图表数据结束行（仅包含学生数据）

    # 创建图表
    create_charts(ws_calc, ws_stat, chart_start_row, chart_end_row)

    # 保存输出文件
    wb.save(output_file)
    print(f"输出文件已保存: {output_file}")


def setup_calc_sheet_headers(ws_calc, ratio_1, ratio_2, ratio_3, bold_font, black_font, center_alignment, thin_border):
    """设置课程目标达成度计算工作表的标题行"""

    # 第一行：配置参数和标题
    # A1-B1: 合并为空
    ws_calc.merge_cells('A1:B1')
    ws_calc.cell(1, 1).border = thin_border
    ws_calc.cell(1, 2).border = thin_border

    # C1: 目标一占比
    ws_calc.cell(1, 3).value = ratio_1
    ws_calc.cell(1, 3).font = black_font
    ws_calc.cell(1, 3).alignment = center_alignment
    ws_calc.cell(1, 3).border = thin_border

    # D1: 目标二占比
    ws_calc.cell(1, 4).value = ratio_2
    ws_calc.cell(1, 4).font = black_font
    ws_calc.cell(1, 4).alignment = center_alignment
    ws_calc.cell(1, 4).border = thin_border

    # E1: 目标三占比
    ws_calc.cell(1, 5).value = ratio_3
    ws_calc.cell(1, 5).font = black_font
    ws_calc.cell(1, 5).alignment = center_alignment
    ws_calc.cell(1, 5).border = thin_border

    # F1-H1: 成绩标题
    ws_calc.merge_cells('F1:H1')
    ws_calc.cell(1, 6).value = '成绩'
    ws_calc.cell(1, 6).font = bold_font
    ws_calc.cell(1, 6).alignment = center_alignment
    ws_calc.cell(1, 6).border = thin_border

    # I1-J1: 合并为空
    ws_calc.merge_cells('I1:J1')
    ws_calc.cell(1, 9).border = thin_border
    ws_calc.cell(1, 10).border = thin_border

    # K1-L1: 平时成绩
    ws_calc.merge_cells('K1:L1')
    ws_calc.cell(1, 11).value = '平时成绩'
    ws_calc.cell(1, 11).font = bold_font
    ws_calc.cell(1, 11).alignment = center_alignment
    ws_calc.cell(1, 11).border = thin_border

    # M1-N1: 平时成绩占比
    ws_calc.merge_cells('M1:N1')
    ws_calc.cell(1, 13).value = REGULAR_SCORE_RATIO
    ws_calc.cell(1, 13).font = black_font
    ws_calc.cell(1, 13).alignment = center_alignment
    ws_calc.cell(1, 13).border = thin_border

    # O1-P1: 期末成绩
    ws_calc.merge_cells('O1:P1')
    ws_calc.cell(1, 15).value = '期末成绩'
    ws_calc.cell(1, 15).font = bold_font
    ws_calc.cell(1, 15).alignment = center_alignment
    ws_calc.cell(1, 15).border = thin_border

    # Q1-R1: 期末成绩占比
    ws_calc.merge_cells('Q1:R1')
    ws_calc.cell(1, 17).value = FINAL_SCORE_RATIO
    ws_calc.cell(1, 17).font = black_font
    ws_calc.cell(1, 17).alignment = center_alignment
    ws_calc.cell(1, 17).border = thin_border

    # S1-V1: 总成绩
    ws_calc.merge_cells('S1:V1')
    ws_calc.cell(1, 19).value = '总成绩'
    ws_calc.cell(1, 19).font = bold_font
    ws_calc.cell(1, 19).alignment = center_alignment
    ws_calc.cell(1, 19).border = thin_border

    # W1-Y1: 达成度
    ws_calc.merge_cells('W1:Y1')
    ws_calc.cell(1, 23).value = '达成度'
    ws_calc.cell(1, 23).font = bold_font
    ws_calc.cell(1, 23).alignment = center_alignment
    ws_calc.cell(1, 23).border = thin_border

    # Z1-AB1: 达成度平均值
    ws_calc.merge_cells('Z1:AB1')
    ws_calc.cell(1, 26).value = '达成度平均值'
    ws_calc.cell(1, 26).font = bold_font
    ws_calc.cell(1, 26).alignment = center_alignment
    ws_calc.cell(1, 26).border = thin_border

    # AC1-AE1: 达成度期望值
    ws_calc.merge_cells('AC1:AE1')
    ws_calc.cell(1, 29).value = '达成度期望值'
    ws_calc.cell(1, 29).font = bold_font
    ws_calc.cell(1, 29).alignment = center_alignment
    ws_calc.cell(1, 29).border = thin_border

    # AF1: 算术平均值
    ws_calc.cell(1, 32).value = '算术平均值'
    ws_calc.cell(1, 32).font = bold_font
    ws_calc.cell(1, 32).alignment = center_alignment
    ws_calc.cell(1, 32).border = thin_border

    # AG1-AG2: 总达成度平均值
    ws_calc.merge_cells('AG1:AG2')
    ws_calc.cell(1, 33).value = '总达成度平均值'
    ws_calc.cell(1, 33).font = bold_font
    ws_calc.cell(1, 33).alignment = center_alignment
    ws_calc.cell(1, 33).border = thin_border
    ws_calc.cell(2, 33).border = thin_border  # 合并单元格的第二行也需要边框

    # 第二行：列标题
    row2_headers = [
        ('A', '班级'), ('B', '学号'), ('C', '目标一'), ('D', '目标二'), ('E', '目标三'),
        ('F', '平时'), ('G', '期末'), ('H', '总分'),
        ('I', '序号'), ('J', '姓名'),
        ('K', '目标1'), ('L', '目标2'), ('M', '目标3'), ('N', '平时'),
        ('O', '目标1'), ('P', '目标2'), ('Q', '目标3'), ('R', '期末'),
        ('S', '目标1'), ('T', '目标2'), ('U', '目标3'), ('V', '总分'),
        ('W', '目标1'), ('X', '目标2'), ('Y', '目标3'),
        ('Z', '目标1'), ('AA', '目标2'), ('AB', '目标3'),
        ('AC', '目标1'), ('AD', '目标2'), ('AE', '目标3'),
        ('AF', '总达成度')
    ]

    for col_letter, header in row2_headers:
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        ws_calc.cell(2, col_idx).value = header
        ws_calc.cell(2, col_idx).font = bold_font
        ws_calc.cell(2, col_idx).alignment = center_alignment
        ws_calc.cell(2, col_idx).border = thin_border


def setup_column_widths(ws_calc):
    """设置列宽"""
    numeric_width = 11

    # 数值列 C-H
    for col in range(3, 9):
        ws_calc.column_dimensions[get_column_letter(col)].width = numeric_width
    # K-Y
    for col in range(11, 26):
        ws_calc.column_dimensions[get_column_letter(col)].width = numeric_width
    # Z-AB
    for col in range(26, 29):
        ws_calc.column_dimensions[get_column_letter(col)].width = numeric_width
    # AC-AE
    for col in range(29, 32):
        ws_calc.column_dimensions[get_column_letter(col)].width = numeric_width

    ws_calc.column_dimensions['AF'].width = numeric_width + 2
    ws_calc.column_dimensions['AG'].width = 16.5
    ws_calc.column_dimensions['A'].width = 13
    ws_calc.column_dimensions['B'].width = 13
    ws_calc.column_dimensions['J'].width = 9
    ws_calc.column_dimensions['I'].width = 6


def setup_statistics_sheet(ws_stat, data_start_row, data_end_row, black_font, bold_font, center_alignment, thin_border):
    """设置达成度统计工作表"""

    # 第一行标题
    ws_stat.cell(1, 1).value = '达成度'
    ws_stat.cell(1, 1).font = bold_font
    ws_stat.cell(1, 1).alignment = center_alignment
    ws_stat.cell(1, 1).border = thin_border

    ws_stat.cell(1, 2).value = '达成情况'
    ws_stat.cell(1, 2).font = bold_font
    ws_stat.cell(1, 2).alignment = center_alignment
    ws_stat.cell(1, 2).border = thin_border

    # C1-D1: 目标1
    ws_stat.merge_cells('C1:D1')
    ws_stat.cell(1, 3).value = '目标1'
    ws_stat.cell(1, 3).font = bold_font
    ws_stat.cell(1, 3).alignment = center_alignment
    ws_stat.cell(1, 3).border = thin_border
    ws_stat.cell(1, 4).border = thin_border  # 合并单元格右侧边框

    # E1-F1: 目标2
    ws_stat.merge_cells('E1:F1')
    ws_stat.cell(1, 5).value = '目标2'
    ws_stat.cell(1, 5).font = bold_font
    ws_stat.cell(1, 5).alignment = center_alignment
    ws_stat.cell(1, 5).border = thin_border
    ws_stat.cell(1, 6).border = thin_border  # 合并单元格右侧边框

    # G1-H1: 目标3
    ws_stat.merge_cells('G1:H1')
    ws_stat.cell(1, 7).value = '目标3'
    ws_stat.cell(1, 7).font = bold_font
    ws_stat.cell(1, 7).alignment = center_alignment
    ws_stat.cell(1, 7).border = thin_border
    ws_stat.cell(1, 8).border = thin_border  # 合并单元格右侧边框

    # 第二行：子标题
    ws_stat.cell(2, 1).border = thin_border
    ws_stat.cell(2, 2).border = thin_border
    ws_stat.cell(2, 3).value = '人数'
    ws_stat.cell(2, 3).font = bold_font
    ws_stat.cell(2, 3).alignment = center_alignment
    ws_stat.cell(2, 3).border = thin_border
    ws_stat.cell(2, 4).value = '占比'
    ws_stat.cell(2, 4).font = bold_font
    ws_stat.cell(2, 4).alignment = center_alignment
    ws_stat.cell(2, 4).border = thin_border
    ws_stat.cell(2, 5).value = '人数'
    ws_stat.cell(2, 5).font = bold_font
    ws_stat.cell(2, 5).alignment = center_alignment
    ws_stat.cell(2, 5).border = thin_border
    ws_stat.cell(2, 6).value = '占比'
    ws_stat.cell(2, 6).font = bold_font
    ws_stat.cell(2, 6).alignment = center_alignment
    ws_stat.cell(2, 6).border = thin_border
    ws_stat.cell(2, 7).value = '人数'
    ws_stat.cell(2, 7).font = bold_font
    ws_stat.cell(2, 7).alignment = center_alignment
    ws_stat.cell(2, 7).border = thin_border
    ws_stat.cell(2, 8).value = '占比'
    ws_stat.cell(2, 8).font = bold_font
    ws_stat.cell(2, 8).alignment = center_alignment
    ws_stat.cell(2, 8).border = thin_border

    # 达成度标准行
    standards = [
        (3, '>0.8', '完全达成'),
        (4, '0.6-0.8', '较好达成'),
        (5, '0.5-0.6', '基本达成'),
        (6, '0.4-0.5', '较少达成'),
        (7, '<0.4', '没有达成')
    ]

    for row, level, desc in standards:
        ws_stat.cell(row, 1).value = level
        ws_stat.cell(row, 1).font = black_font
        ws_stat.cell(row, 1).alignment = center_alignment
        ws_stat.cell(row, 1).border = thin_border

        ws_stat.cell(row, 2).value = desc
        ws_stat.cell(row, 2).font = black_font
        ws_stat.cell(row, 2).alignment = center_alignment
        ws_stat.cell(row, 2).border = thin_border

    # 人数统计公式
    # 目标1 (W列)
    ws_stat.cell(3, 3).value = f'=COUNTIF(\'课程目标达成度计算\'!W{data_start_row}:W{data_end_row},">0.8")'
    ws_stat.cell(4, 3).value = f'=COUNTIFS(\'课程目标达成度计算\'!W{data_start_row}:W{data_end_row},">=0.6",\'课程目标达成度计算\'!W{data_start_row}:W{data_end_row},"<=0.8")'
    ws_stat.cell(5, 3).value = f'=COUNTIFS(\'课程目标达成度计算\'!W{data_start_row}:W{data_end_row},">=0.5",\'课程目标达成度计算\'!W{data_start_row}:W{data_end_row},"<0.6")'
    ws_stat.cell(6, 3).value = f'=COUNTIFS(\'课程目标达成度计算\'!W{data_start_row}:W{data_end_row},">=0.4",\'课程目标达成度计算\'!W{data_start_row}:W{data_end_row},"<0.5")'
    ws_stat.cell(7, 3).value = f'=COUNTIF(\'课程目标达成度计算\'!W{data_start_row}:W{data_end_row},"<0.4")'

    # 目标2 (X列)
    ws_stat.cell(3, 5).value = f'=COUNTIF(\'课程目标达成度计算\'!X{data_start_row}:X{data_end_row},">0.8")'
    ws_stat.cell(4, 5).value = f'=COUNTIFS(\'课程目标达成度计算\'!X{data_start_row}:X{data_end_row},">=0.6",\'课程目标达成度计算\'!X{data_start_row}:X{data_end_row},"<=0.8")'
    ws_stat.cell(5, 5).value = f'=COUNTIFS(\'课程目标达成度计算\'!X{data_start_row}:X{data_end_row},">=0.5",\'课程目标达成度计算\'!X{data_start_row}:X{data_end_row},"<0.6")'
    ws_stat.cell(6, 5).value = f'=COUNTIFS(\'课程目标达成度计算\'!X{data_start_row}:X{data_end_row},">=0.4",\'课程目标达成度计算\'!X{data_start_row}:X{data_end_row},"<0.5")'
    ws_stat.cell(7, 5).value = f'=COUNTIF(\'课程目标达成度计算\'!X{data_start_row}:X{data_end_row},"<0.4")'

    # 目标3 (Y列)
    ws_stat.cell(3, 7).value = f'=COUNTIF(\'课程目标达成度计算\'!Y{data_start_row}:Y{data_end_row},">0.8")'
    ws_stat.cell(4, 7).value = f'=COUNTIFS(\'课程目标达成度计算\'!Y{data_start_row}:Y{data_end_row},">=0.6",\'课程目标达成度计算\'!Y{data_start_row}:Y{data_end_row},"<=0.8")'
    ws_stat.cell(5, 7).value = f'=COUNTIFS(\'课程目标达成度计算\'!Y{data_start_row}:Y{data_end_row},">=0.5",\'课程目标达成度计算\'!Y{data_start_row}:Y{data_end_row},"<0.6")'
    ws_stat.cell(6, 7).value = f'=COUNTIFS(\'课程目标达成度计算\'!Y{data_start_row}:Y{data_end_row},">=0.4",\'课程目标达成度计算\'!Y{data_start_row}:Y{data_end_row},"<0.5")'
    ws_stat.cell(7, 7).value = f'=COUNTIF(\'课程目标达成度计算\'!Y{data_start_row}:Y{data_end_row},"<0.4")'

    # 占比公式和样式
    for row in range(3, 8):
        # 人数列样式
        for col in [3, 5, 7]:
            ws_stat.cell(row, col).font = black_font
            ws_stat.cell(row, col).alignment = center_alignment
            ws_stat.cell(row, col).border = thin_border

        # 占比公式 - 使用 COUNT() 统计有效学生数（排除空值）
        ws_stat.cell(row, 4).value = f'=C{row}/COUNT(\'课程目标达成度计算\'!W${data_start_row}:W${data_end_row})'
        ws_stat.cell(row, 4).font = black_font
        ws_stat.cell(row, 4).alignment = center_alignment
        ws_stat.cell(row, 4).border = thin_border
        ws_stat.cell(row, 4).number_format = '0.00%'

        ws_stat.cell(row, 6).value = f'=E{row}/COUNT(\'课程目标达成度计算\'!X${data_start_row}:X${data_end_row})'
        ws_stat.cell(row, 6).font = black_font
        ws_stat.cell(row, 6).alignment = center_alignment
        ws_stat.cell(row, 6).border = thin_border
        ws_stat.cell(row, 6).number_format = '0.00%'

        ws_stat.cell(row, 8).value = f'=G{row}/COUNT(\'课程目标达成度计算\'!Y${data_start_row}:Y${data_end_row})'
        ws_stat.cell(row, 8).font = black_font
        ws_stat.cell(row, 8).alignment = center_alignment
        ws_stat.cell(row, 8).border = thin_border
        ws_stat.cell(row, 8).number_format = '0.00%'

    # 设置列宽
    ws_stat.column_dimensions['A'].width = 11
    ws_stat.column_dimensions['B'].width = 11
    for col in ['C', 'D', 'E', 'F', 'G', 'H']:
        ws_stat.column_dimensions[col].width = 7


def create_charts(ws_calc, ws_stat, data_start_row, data_end_row):
    """创建所有图表"""
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.marker import Marker

    # ==================== 课程目标达成度计算页的折线图 ====================
    chart_configs = [
        {
            'title': '目标1达成度',
            'y_col': 23,  # W列 - 达成度目标1
            'avg_col': 26,  # Z列 - 平均值
            'exp_col': 29,  # AC列 - 期望值
        },
        {
            'title': '目标2达成度',
            'y_col': 24,  # X列
            'avg_col': 27,  # AA列
            'exp_col': 30,  # AD列
        },
        {
            'title': '目标3达成度',
            'y_col': 25,  # Y列
            'avg_col': 28,  # AB列
            'exp_col': 31,  # AE列
        },
        {
            'title': '总达成度',
            'y_col': 32,  # AF列
            'avg_col': 33,  # AG列
            'exp_col': 29,  # AC列 - 使用同一个期望值
        },
    ]

    chart_width = 18
    chart_height = 12
    start_col = 36  # AJ列
    col_gap = 12  # 右侧图表与AV列对齐
    row_gap = 24
    row1_start = 2

    for i, config in enumerate(chart_configs):
        chart = LineChart()
        chart.title = config['title']
        chart.style = 10
        chart.x_axis.title = '学生序号'
        chart.y_axis.title = '达成度'
        chart.legend = None  # 隐藏图例

        # 设置Y轴范围
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 1

        # 设置网格线（X轴和Y轴）
        from openpyxl.chart.axis import ChartLines
        from openpyxl.chart.shapes import GraphicalProperties
        from openpyxl.drawing.line import LineProperties
        # 使用浅灰色模拟透明效果（#C0C0C0 约等于 60% 透明的黑色）
        gridline_props = GraphicalProperties(
            ln=LineProperties(solidFill='C0C0C0', w=9525)  # 0.75pt 线宽
        )
        chart.x_axis.majorGridlines = ChartLines(spPr=gridline_props)
        chart.y_axis.majorGridlines = ChartLines(spPr=gridline_props)

        # 设置X轴刻度间隔（分类轴使用 tickLblSkip）
        chart.x_axis.tickLblSkip = 5  # 每隔5个显示一个标签
        chart.x_axis.tickMarkSkip = 5  # 每隔5个显示一个刻度线

        # X轴数据（I列，序号从1开始）
        x_values = Reference(ws_calc, min_col=9, min_row=data_start_row, max_row=data_end_row)

        # 系列1: 达成度数据点
        y_values = Reference(ws_calc, min_col=config['y_col'], min_row=data_start_row - 1, max_row=data_end_row)
        chart.add_data(y_values, titles_from_data=True)

        # 系列2: 平均值线
        avg_values = Reference(ws_calc, min_col=config['avg_col'], min_row=data_start_row - 1, max_row=data_end_row)
        chart.add_data(avg_values, titles_from_data=True)

        # 系列3: 期望值线
        exp_values = Reference(ws_calc, min_col=config['exp_col'], min_row=data_start_row - 1, max_row=data_end_row)
        chart.add_data(exp_values, titles_from_data=True)

        # 设置X轴分类
        chart.set_categories(x_values)

        # 设置系列样式
        if len(chart.series) >= 1:
            # 系列1: 只有标记点，没有连线
            chart.series[0].marker = Marker(symbol='circle', size=5)
            chart.series[0].graphicalProperties.line = LineProperties(noFill=True)

        if len(chart.series) >= 2:
            # 系列2: 鲜绿色双线+系统点线（平均值）
            chart.series[1].marker = Marker(symbol='none')
            chart.series[1].graphicalProperties.line = LineProperties(
                solidFill='00FF00',
                w=25000,
                cmpd='dbl',
                prstDash='sysDot'
            )

        if len(chart.series) >= 3:
            # 系列3: 红色双线+系统点线（期望值）
            chart.series[2].marker = Marker(symbol='none')
            chart.series[2].graphicalProperties.line = LineProperties(
                solidFill='FF0000',
                w=25000,
                cmpd='dbl',
                prstDash='sysDot'
            )

        # 设置位置和尺寸
        col_offset = (i % 2) * col_gap
        row_offset = (i // 2) * row_gap
        chart.anchor = f'{get_column_letter(start_col + col_offset)}{row1_start + row_offset}'
        chart.width = chart_width
        chart.height = chart_height

        ws_calc.add_chart(chart)

    print("散点图已创建")

    # ==================== 达成度统计页的柱状图 ====================
    stat_chart_configs = [
        {'title': '目标1达成度人数占比统计', 'data_col': 4, 'anchor_col': 1},   # D列数据，A列位置
        {'title': '目标2达成度人数占比统计', 'data_col': 6, 'anchor_col': 9},   # F列数据，I列位置
        {'title': '目标3达成度人数占比统计', 'data_col': 8, 'anchor_col': 16},  # H列数据，P列位置
    ]

    stat_chart_width = 10
    stat_chart_height = 10
    stat_start_row = 9

    for i, config in enumerate(stat_chart_configs):
        chart = BarChart()
        chart.title = config['title']
        chart.style = 10
        chart.type = 'col'
        chart.grouping = 'clustered'
        chart.legend = None  # 去除图例

        # 设置Y轴范围 0%-100%
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 1
        chart.y_axis.numFmt = '0%'  # 百分比格式

        # 数据范围
        data = Reference(ws_stat, min_col=config['data_col'], min_row=2, max_row=7)
        cats = Reference(ws_stat, min_col=2, min_row=3, max_row=7)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # X轴标签不旋转
        from openpyxl.chart.text import RichText
        from openpyxl.drawing.text import RichTextProperties, Paragraph, ParagraphProperties, CharacterProperties
        chart.x_axis.txPr = RichText(
            bodyPr=RichTextProperties(rot=0),
            p=[Paragraph(
                pPr=ParagraphProperties(
                    defRPr=CharacterProperties(sz=900)
                )
            )]
        )

        # 设置位置和尺寸
        chart.anchor = f'{get_column_letter(config["anchor_col"])}{stat_start_row}'
        chart.width = stat_chart_width
        chart.height = stat_chart_height

        ws_stat.add_chart(chart)

    print("柱状图已创建")


def process_single_file(grades_file, output_file):
    """处理单个成绩文件"""
    print(f"\n处理文件: {grades_file}")

    # 1. 提取学生数据
    print("  [1/3] 从成绩文件提取学生数据...")
    students = extract_students_from_grades(grades_file)
    print(f"  成功提取 {len(students)} 名学生数据")

    # 2. 排序
    print("  [2/3] 按行政班分组，按学号升序排序...")
    students = sort_students(students)

    # 显示排序后的班级统计
    from collections import Counter
    classes = Counter([s['class'] for s in students])
    for cls, count in sorted(classes.items()):
        print(f"    {cls}: {count}人")

    # 3. 创建工作簿并输出
    print("  [3/3] 创建工作簿...")
    create_workbook(output_file, students)
    print(f"  输出文件: {output_file}")


def batch_process():
    """批处理模式：遍历成绩单目录，生成达成度报告"""
    import os

    # 目录配置
    input_dir = '/Users/zhiqiliu/Documents/百度网盘同步空间/Python_Projects_Sync/达成度报告Excel制作/成绩单'
    output_dir = '/Users/zhiqiliu/Documents/百度网盘同步空间/Python_Projects_Sync/达成度报告Excel制作/达成度数据输出'

    print("=" * 50)
    print("达成度数据处理脚本（批处理模式）")
    print("=" * 50)
    print(f"\n输入目录: {input_dir}")
    print(f"输出目录: {output_dir}")

    # 确保输出目录存在
    os.makedirs(output_dir, exist_ok=True)

    # 获取所有 Excel 文件
    excel_files = [f for f in os.listdir(input_dir)
                   if f.endswith(('.xlsx', '.xls')) and not f.startswith(('.', '~$'))]

    if not excel_files:
        print("\n未找到 Excel 文件！")
        return

    print(f"\n找到 {len(excel_files)} 个 Excel 文件:")
    for f in excel_files:
        print(f"  - {f}")

    # 处理每个文件
    success_count = 0
    fail_count = 0

    for filename in excel_files:
        input_path = os.path.join(input_dir, filename)

        # 生成输出文件名：原文件名_达成度报告.xlsx
        name_without_ext = os.path.splitext(filename)[0]
        output_filename = f"{name_without_ext}_达成度报告.xlsx"
        output_path = os.path.join(output_dir, output_filename)

        try:
            process_single_file(input_path, output_path)
            success_count += 1
        except Exception as e:
            print(f"\n  错误: 处理 {filename} 失败 - {e}")
            fail_count += 1

    # 汇总
    print("\n" + "=" * 50)
    print(f"批处理完成！成功: {success_count}, 失败: {fail_count}")
    print("=" * 50)


def main():
    """主函数"""
    import sys

    # 检查命令行参数
    if len(sys.argv) > 1 and sys.argv[1] == '--batch':
        batch_process()
    else:
        # 单文件处理模式（向后兼容）
        grades_file = '2022-2023第一学期总评成绩(按行政班).xlsx'
        output_file = '达成度数据输出.xlsx'

        print("=" * 50)
        print("达成度数据处理脚本（独立版）")
        print("=" * 50)
        print("\n提示: 使用 --batch 参数启用批处理模式")

        # 1. 提取学生数据
        print("\n[1/3] 从成绩文件提取学生数据...")
        students = extract_students_from_grades(grades_file)
        print(f"成功提取 {len(students)} 名学生数据")

        # 2. 排序
        print("\n[2/3] 按行政班分组，按学号升序排序...")
        students = sort_students(students)

        # 显示排序后的班级统计
        from collections import Counter
        classes = Counter([s['class'] for s in students])
        for cls, count in sorted(classes.items()):
            print(f"  {cls}: {count}人")

        # 3. 创建工作簿并输出
        print("\n[3/3] 创建工作簿...")
        create_workbook(output_file, students)

        print("\n" + "=" * 50)
        print("处理完成！")
        print("=" * 50)


if __name__ == '__main__':
    main()
